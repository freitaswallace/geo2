#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificador de Consistência de Documentos de Georreferenciamento
Aplicação GUI para cartórios - Análise multimodal com Gemini AI
Autor: Sistema Automatizado
Versão: 4.0 - Interface moderna com Modo Automático
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from pathlib import Path
import threading
from typing import List, Optional, Dict, Tuple
import json
import tempfile
import shutil
import webbrowser
import math
from datetime import datetime
import configparser
import time


def get_poppler_path():
    """
    Detecta o caminho do Poppler, seja rodando como script ou executável.

    Quando compilado com PyInstaller, o Poppler está em sys._MEIPASS/poppler/bin
    Quando rodando como script, retorna None (usa instalação do sistema)
    """
    if getattr(sys, 'frozen', False):
        # Rodando como executável PyInstaller
        base_path = Path(sys._MEIPASS)
        poppler_path = base_path / 'poppler' / 'bin'
        if poppler_path.exists():
            print(f"✅ Poppler embutido encontrado em: {poppler_path}")
            return str(poppler_path)
        else:
            print(f"⚠️ Poppler não encontrado em: {poppler_path}")
            return None
    else:
        # Rodando como script Python
        print("ℹ️ Usando Poppler do sistema")
        return None


# Detectar caminho do Poppler globalmente
POPPLER_PATH = get_poppler_path()

try:
    from pdf2image import convert_from_path
    from PIL import Image, ImageTk
    import google.generativeai as genai
    from openpyxl import load_workbook
    import PyPDF2
    # Importar funções de extração do script existente
    from process_memorial_descritivo_v2 import (
        extract_table_from_pdf,
        extrair_memorial_incra,
        create_excel_file
    )
except ImportError as e:
    print(f"❌ Erro: Biblioteca necessária não encontrada: {e}")
    print("\nInstale as dependências com:")
    print("pip install pdf2image Pillow google-generativeai openpyxl PyPDF2")
    print("\nNota: Também é necessário ter o 'poppler-utils' instalado no sistema.")
    sys.exit(1)


class ConfigManager:
    """Gerencia configurações persistentes da aplicação."""

    def __init__(self):
        self.config_dir = Path.home() / ".conferencia_geo"
        self.config_file = self.config_dir / "config.ini"
        self.config = configparser.ConfigParser()
        self._ensure_config_exists()

    def _ensure_config_exists(self):
        """Cria diretório e arquivo de configuração se não existir."""
        self.config_dir.mkdir(parents=True, exist_ok=True)
        if not self.config_file.exists():
            self.config['API'] = {'gemini_key': ''}
            self.save()
        else:
            self.config.read(self.config_file)

    def save(self):
        """Salva configurações no arquivo."""
        with open(self.config_file, 'w') as f:
            self.config.write(f)

    def get_api_key(self) -> str:
        """Retorna a API key salva."""
        return self.config.get('API', 'gemini_key', fallback='')

    def set_api_key(self, key: str):
        """Salva a API key."""
        if 'API' not in self.config:
            self.config['API'] = {}
        self.config['API']['gemini_key'] = key
        self.save()


class VerificadorGeorreferenciamento:
    """Classe principal da aplicação de verificação de documentos."""

    def __init__(self, root):
        self.root = root
        self.root.title("✨ Verificador INCRA Pro v4.0")

        # Maximizar janela ao iniciar
        try:
            self.root.state('zoomed')  # Windows
        except:
            try:
                self.root.attributes('-zoomed', True)  # Linux
            except:
                self.root.geometry("1450x980")  # Fallback

        # Configurar handler para fechamento da janela
        self.root.protocol("WM_DELETE_WINDOW", self._ao_fechar_aplicacao)

        # Gerenciador de configurações
        self.config_manager = ConfigManager()

        # Variáveis para armazenar caminhos dos arquivos
        self.incra_path = tk.StringVar()
        self.projeto_path = tk.StringVar()
        self.numero_prenotacao = tk.StringVar()
        self.modo_atual = tk.StringVar(value="automatico")

        # Variáveis para sub-modo automático
        self.modo_automatico_tipo = tk.StringVar(value="paginas")  # "ia" ou "paginas"
        self.paginas_incra_auto = tk.StringVar()  # Para modo automático
        self.paginas_projeto_auto = tk.StringVar()  # Para modo automático
        self.arquivo_manual_incra_auto = tk.StringVar()  # Arquivo INCRA selecionado manualmente (modo auto)
        self.arquivo_manual_projeto_auto = tk.StringVar()  # Arquivo Projeto selecionado manualmente (modo auto)

        # Variáveis para sub-modo manual
        self.modo_manual_tipo = tk.StringVar(value="completo")  # "completo" ou "por_paginas"
        self.paginas_incra_manual = tk.StringVar()  # Para modo manual por páginas
        self.paginas_projeto_manual = tk.StringVar()  # Para modo manual por páginas

        # Variáveis para armazenar dados extraídos
        self.incra_excel_path: Optional[str] = None
        self.projeto_excel_path: Optional[str] = None
        self.incra_data: Optional[Dict] = None
        self.projeto_data: Optional[Dict] = None

        # Variáveis para modo automático
        self.pdf_extraido_incra: Optional[str] = None
        self.pdf_extraido_projeto: Optional[str] = None
        self.preview_incra_image: Optional[Image.Image] = None
        self.preview_projeto_image: Optional[Image.Image] = None

        # Variáveis para janela de progresso
        self.progress_window = None
        self.progress_bar = None
        self.progress_label = None
        self.progress_detail_label = None

        # Configurar estilo moderno
        self._configurar_estilo()

        # Criar interface
        self._criar_interface()

        # Carregar API key salva
        self._carregar_api_key()

    def _configurar_estilo(self):
        """Configura tema moderno e profissional com cores vibrantes."""
        style = ttk.Style()
        style.theme_use('clam')

        # Paleta de cores moderna e agradável (inspirada em Material Design)
        self.colors = {
            'primary': '#6366F1',      # Indigo vibrante
            'primary_dark': '#4F46E5',
            'secondary': '#EC4899',    # Rosa vibrante
            'success': '#10B981',      # Verde esmeralda
            'warning': '#F59E0B',      # Âmbar
            'danger': '#EF4444',       # Vermelho
            'info': '#3B82F6',         # Azul
            'bg_light': '#F9FAFB',     # Cinza muito claro
            'bg_card': '#FFFFFF',
            'text_dark': '#1F2937',
            'text_medium': '#6B7280',
            'text_light': '#9CA3AF',
            'border': '#E5E7EB'
        }

        # Configurar background
        self.root.configure(bg=self.colors['bg_light'])

        # Estilos de labels
        style.configure('Title.TLabel',
            font=('Inter', 24, 'bold'),
            foreground=self.colors['primary'],
            background=self.colors['bg_light']
        )

        style.configure('Subtitle.TLabel',
            font=('Inter', 13, 'bold'),
            foreground=self.colors['text_dark'],
            background=self.colors['bg_light']
        )

        style.configure('Normal.TLabel',
            font=('Inter', 10),
            foreground=self.colors['text_medium'],
            background=self.colors['bg_light']
        )

        style.configure('Emoji.TLabel',
            font=('Segoe UI Emoji', 32),
            background=self.colors['bg_card']
        )

        # Estilos de botões
        style.configure('Primary.TButton',
            font=('Inter', 12, 'bold'),
            padding=(20, 15),
            borderwidth=0
        )

        style.map('Primary.TButton',
            background=[('active', self.colors['primary_dark']), ('!active', self.colors['primary'])],
            foreground=[('active', 'white'), ('!active', 'white')]
        )

        style.configure('Success.TButton',
            font=('Inter', 11, 'bold'),
            padding=(15, 12)
        )

        style.configure('Action.TButton',
            font=('Inter', 10, 'bold'),
            padding=(10, 8)
        )

        # Estilos de frames
        style.configure('Card.TFrame',
            background=self.colors['bg_card'],
            relief='flat'
        )

        style.configure('TFrame',
            background=self.colors['bg_light']
        )

        # Estilos de LabelFrame
        style.configure('Card.TLabelframe',
            background=self.colors['bg_card'],
            borderwidth=0
        )

        style.configure('Card.TLabelframe.Label',
            font=('Inter', 12, 'bold'),
            foreground=self.colors['primary'],
            background=self.colors['bg_card']
        )

    def _criar_interface(self):
        """Cria todos os elementos da interface gráfica."""

        # Container principal com scrollbar
        container = tk.Frame(self.root, bg=self.colors['bg_light'])
        container.pack(fill=tk.BOTH, expand=True)

        # Canvas e Scrollbar
        canvas = tk.Canvas(container, bg=self.colors['bg_light'], highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)

        # Frame scrollável dentro do canvas
        main_frame = tk.Frame(canvas, bg=self.colors['bg_light'])

        # Posicionar scrollbar e canvas
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Criar window no canvas
        canvas_window = canvas.create_window((0, 0), window=main_frame, anchor="nw")

        # Configurar scroll
        canvas.configure(yscrollcommand=scrollbar.set)

        # Atualizar região de scroll quando o conteúdo mudar
        def configure_scroll(event=None):
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Ajustar largura do main_frame para preencher o canvas
            canvas_width = canvas.winfo_width()
            if canvas_width > 1:  # Só atualizar se o canvas tiver largura válida
                canvas.itemconfig(canvas_window, width=canvas_width)

        main_frame.bind("<Configure>", configure_scroll)
        canvas.bind("<Configure>", configure_scroll)

        # Scroll com mouse wheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # Forçar atualização inicial após 100ms
        self.root.after(100, configure_scroll)

        # ===== CABEÇALHO COM DESIGN MODERNO =====
        header_frame = tk.Frame(main_frame, bg=self.colors['bg_light'])
        header_frame.pack(fill=tk.X, pady=(0, 25), padx=25)

        # Título com emoji grande
        title_container = tk.Frame(header_frame, bg=self.colors['bg_light'])
        title_container.pack()

        tk.Label(
            title_container,
            text="🏛️",
            font=('Segoe UI Emoji', 48),
            bg=self.colors['bg_light']
        ).pack(side=tk.LEFT, padx=(0, 15))

        title_text_frame = tk.Frame(title_container, bg=self.colors['bg_light'])
        title_text_frame.pack(side=tk.LEFT)

        ttk.Label(
            title_text_frame,
            text="VERIFICADOR INCRA PRO",
            style='Title.TLabel'
        ).pack(anchor=tk.W)

        ttk.Label(
            title_text_frame,
            text="Sistema Inteligente de Análise e Conferência Georreferenciada",
            style='Normal.TLabel'
        ).pack(anchor=tk.W)

        # ===== BARRA DE FERRAMENTAS COM CARDS =====
        # Card com borda destacada para ferramentas principais
        toolbar_card = tk.Frame(
            main_frame,
            bg=self.colors['bg_card'],
            highlightbackground=self.colors['primary'],
            highlightthickness=2,
            relief=tk.SOLID,
            borderwidth=1
        )
        toolbar_card.pack(fill=tk.X, pady=(0, 20), padx=25)

        toolbar_content = tk.Frame(toolbar_card, bg=self.colors['bg_card'])
        toolbar_content.pack(fill=tk.X, padx=20, pady=15)

        # Botão API Key estilizado
        api_frame = tk.Frame(toolbar_content, bg=self.colors['bg_card'])
        api_frame.pack(side=tk.LEFT, padx=(0, 20))

        tk.Button(
            api_frame,
            text="⚙️  Configurar API",
            command=self._abrir_config_api,
            font=('Inter', 10, 'bold'),
            bg=self.colors['info'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor='hand2',
            activebackground=self.colors['primary'],
            highlightthickness=2,
            highlightbackground=self.colors['info'],
            highlightcolor=self.colors['primary_dark']
        ).pack()

        # Status API
        self.api_status_label = tk.Label(
            api_frame,
            text="⭕ Não configurada",
            font=('Inter', 8),
            fg=self.colors['danger'],
            bg=self.colors['bg_card']
        )
        self.api_status_label.pack(pady=(5, 0))

        # Separador vertical
        tk.Frame(
            toolbar_content,
            width=2,
            bg=self.colors['border']
        ).pack(side=tk.LEFT, fill=tk.Y, padx=20)

        # Campo Prenotação estilizado
        prenotacao_frame = tk.Frame(toolbar_content, bg=self.colors['bg_card'])
        prenotacao_frame.pack(side=tk.LEFT)

        tk.Label(
            prenotacao_frame,
            text="📋",
            font=('Segoe UI Emoji', 20),
            bg=self.colors['bg_card']
        ).pack(side=tk.LEFT, padx=(0, 10))

        prenotacao_input_frame = tk.Frame(prenotacao_frame, bg=self.colors['bg_card'])
        prenotacao_input_frame.pack(side=tk.LEFT)

        tk.Label(
            prenotacao_input_frame,
            text="Nº Prenotação",
            font=('Inter', 11, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(anchor=tk.W)

        prenotacao_entry = tk.Entry(
            prenotacao_input_frame,
            textvariable=self.numero_prenotacao,
            font=('Inter', 13, 'bold'),
            width=15,
            relief=tk.SOLID,
            bg='#F3F4F6',
            fg=self.colors['primary'],
            insertbackground=self.colors['primary'],
            borderwidth=2,
            highlightthickness=0
        )
        prenotacao_entry.pack(pady=(5, 0), ipady=6, ipadx=8)

        vcmd = (self.root.register(self._validar_numero), '%P')
        prenotacao_entry.config(validate='key', validatecommand=vcmd)

        # Separador vertical
        tk.Frame(
            toolbar_content,
            width=2,
            bg=self.colors['border']
        ).pack(side=tk.LEFT, fill=tk.Y, padx=20)

        # Botão Limpar Backups (redesenhado)
        limpar_frame = tk.Frame(toolbar_content, bg=self.colors['bg_card'])
        limpar_frame.pack(side=tk.LEFT, padx=(0, 10))

        # Container para centralizar conteúdo
        limpar_content = tk.Frame(limpar_frame, bg=self.colors['bg_card'])
        limpar_content.pack()

        # Ícone e botão em um frame
        btn_container = tk.Frame(limpar_content, bg=self.colors['bg_card'])
        btn_container.pack()

        # Ícone decorativo
        tk.Label(
            btn_container,
            text="🗑️",
            font=('Segoe UI Emoji', 16),
            bg=self.colors['bg_card']
        ).pack(side=tk.LEFT, padx=(0, 8))

        # Botão com design melhorado
        limpar_btn = tk.Button(
            btn_container,
            text="Limpar Backups",
            command=self._limpar_arquivos_backup,
            font=('Inter', 9, 'bold'),
            bg='#DC2626',
            fg='white',
            relief=tk.FLAT,
            padx=12,
            pady=6,
            cursor='hand2',
            activebackground='#B91C1C',
            highlightthickness=0
        )
        limpar_btn.pack(side=tk.LEFT)

        # Efeito hover
        def on_enter(e):
            limpar_btn.config(bg='#B91C1C')

        def on_leave(e):
            limpar_btn.config(bg='#DC2626')

        limpar_btn.bind('<Enter>', on_enter)
        limpar_btn.bind('<Leave>', on_leave)

        # Status de limpeza (mais discreto)
        self.backup_status_label = tk.Label(
            limpar_content,
            text="",
            font=('Inter', 7),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card']
        )
        self.backup_status_label.pack(pady=(3, 0))

        # ===== SELETOR DE MODO (CARDS GRANDES E BONITOS) =====
        modo_card = self._criar_card(main_frame)
        modo_card.pack(fill=tk.X, pady=(0, 20), padx=25)

        modo_content = tk.Frame(modo_card, bg=self.colors['bg_card'])
        modo_content.pack(fill=tk.X, padx=20, pady=20)

        tk.Label(
            modo_content,
            text="Escolha o modo de operação:",
            font=('Inter', 13, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 15))

        # Container para os cards de modo
        modos_container = tk.Frame(modo_content, bg=self.colors['bg_card'])
        modos_container.pack(fill=tk.X)

        # CARD MODO AUTOMÁTICO
        self.card_automatico = self._criar_modo_card(
            modos_container,
            "🤖",
            "MODO AUTOMÁTICO",
            "Busca inteligente na rede\nExtração automática com IA\nMais rápido e eficiente",
            self.colors['primary'],
            lambda: self._selecionar_modo("automatico")
        )
        self.card_automatico.pack(side=tk.LEFT, padx=(0, 15), expand=True, fill=tk.BOTH)

        # CARD MODO MANUAL
        self.card_manual = self._criar_modo_card(
            modos_container,
            "📝",
            "MODO MANUAL",
            "Selecione os arquivos manualmente\nMaior controle sobre os documentos\nRecomendado para casos especiais",
            self.colors['secondary'],
            lambda: self._selecionar_modo("manual")
        )
        self.card_manual.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        # ===== CONTEÚDO DO MODO SELECIONADO =====
        self.content_frame = tk.Frame(main_frame, bg=self.colors['bg_light'])
        self.content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20), padx=25)

        # Criar ambos os modos (esconder um deles)
        self._criar_modo_automatico_content()
        self._criar_modo_manual_content()

        # Selecionar modo inicial
        self._selecionar_modo("automatico")

        # ===== ÁREA DE RESULTADOS =====
        result_card = self._criar_card(main_frame)
        result_card.pack(fill=tk.BOTH, expand=True, padx=25)

        result_content = tk.Frame(result_card, bg=self.colors['bg_card'])
        result_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        tk.Label(
            result_content,
            text="📊  Relatório de Comparação",
            font=('Inter', 12, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['bg_card']
        ).pack(anchor=tk.W, pady=(0, 10))

        # ScrolledText com estilo
        self.resultado_text = scrolledtext.ScrolledText(
            result_content,
            font=('Consolas', 10),
            wrap=tk.WORD,
            relief=tk.SOLID,
            bg='#F9FAFB',
            fg=self.colors['text_dark'],
            insertbackground=self.colors['primary'],
            selectbackground=self.colors['primary'],
            selectforeground='white',
            borderwidth=2,
            highlightthickness=0
        )
        self.resultado_text.pack(fill=tk.BOTH, expand=True, ipady=10, ipadx=10)

        # Botão Limpar (para nova conferência)
        limpar_btn_frame = tk.Frame(result_content, bg=self.colors['bg_card'])
        limpar_btn_frame.pack(pady=(15, 0))

        self.btn_limpar = tk.Button(
            limpar_btn_frame,
            text="🔄  LIMPAR E FAZER NOVA CONFERÊNCIA",
            command=self._limpar_dados,
            font=('Inter', 11, 'bold'),
            bg='#F59E0B',
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=12,
            cursor='hand2',
            activebackground='#D97706',
            activeforeground='white'
        )
        self.btn_limpar.pack()

        # Hover effect
        def on_enter_limpar(e):
            self.btn_limpar.config(bg='#D97706')
        def on_leave_limpar(e):
            self.btn_limpar.config(bg='#F59E0B')

        self.btn_limpar.bind('<Enter>', on_enter_limpar)
        self.btn_limpar.bind('<Leave>', on_leave_limpar)

        # ===== BARRA DE STATUS =====
        status_frame = tk.Frame(main_frame, bg=self.colors['bg_card'], height=40)
        status_frame.pack(fill=tk.X, pady=(15, 25), padx=25)

        self.status_label = tk.Label(
            status_frame,
            text="✨ Pronto para iniciar",
            font=('Inter', 10),
            fg=self.colors['success'],
            bg=self.colors['bg_card'],
            anchor=tk.W
        )
        self.status_label.pack(fill=tk.X, padx=15, pady=10)

    def _criar_card(self, parent):
        """Cria um card (frame com sombra e bordas arredondadas simuladas)."""
        card = tk.Frame(
            parent,
            bg=self.colors['bg_card'],
            highlightbackground=self.colors['border'],
            highlightthickness=1
        )
        return card

    def _criar_modo_card(self, parent, emoji, titulo, descricao, cor, comando):
        """Cria um card clicável para seleção de modo."""
        card = tk.Frame(
            parent,
            bg=self.colors['bg_card'],
            highlightbackground=self.colors['border'],
            highlightthickness=2,
            cursor='hand2'
        )

        # Conteúdo interno
        content = tk.Frame(card, bg=self.colors['bg_card'])
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)

        # Emoji grande
        tk.Label(
            content,
            text=emoji,
            font=('Segoe UI Emoji', 48),
            bg=self.colors['bg_card']
        ).pack(pady=(0, 15))

        # Título
        tk.Label(
            content,
            text=titulo,
            font=('Inter', 14, 'bold'),
            fg=cor,
            bg=self.colors['bg_card']
        ).pack()

        # Descrição
        tk.Label(
            content,
            text=descricao,
            font=('Inter', 9),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card'],
            justify=tk.CENTER
        ).pack(pady=(10, 0))

        # Badge de status (inicialmente oculto)
        badge = tk.Label(
            content,
            text="✓ SELECIONADO",
            font=('Inter', 8, 'bold'),
            fg='white',
            bg=cor,
            padx=10,
            pady=4
        )

        # Evento de clique
        def on_click(event=None):
            comando()

        card.bind('<Button-1>', on_click)
        for widget in card.winfo_children():
            widget.bind('<Button-1>', on_click)
            for child in widget.winfo_children():
                child.bind('<Button-1>', on_click)

        # Guardar referências para atualização
        card.badge = badge
        card.cor = cor
        card.content_frame = content

        return card

    def _selecionar_modo(self, modo):
        """Alterna entre modos e atualiza visual dos cards."""
        self.modo_atual.set(modo)

        # Atualizar visual dos cards
        if modo == "automatico":
            # Destacar automático
            self.card_automatico.config(highlightbackground=self.colors['primary'], highlightthickness=3)
            self.card_automatico.badge.pack(pady=(15, 0))

            # Desmarcar manual
            self.card_manual.config(highlightbackground=self.colors['border'], highlightthickness=2)
            self.card_manual.badge.pack_forget()

            # Mostrar conteúdo
            self.manual_content.pack_forget()
            self.automatico_content.pack(fill=tk.BOTH, expand=True)

        else:  # manual
            # Destacar manual
            self.card_manual.config(highlightbackground=self.colors['secondary'], highlightthickness=3)
            self.card_manual.badge.pack(pady=(15, 0))

            # Desmarcar automático
            self.card_automatico.config(highlightbackground=self.colors['border'], highlightthickness=2)
            self.card_automatico.badge.pack_forget()

            # Mostrar conteúdo
            self.automatico_content.pack_forget()
            self.manual_content.pack(fill=tk.BOTH, expand=True)

    def _alternar_modo_automatico(self):
        """Alterna entre sub-modos do modo automático (IA vs Páginas)."""
        modo = self.modo_automatico_tipo.get()

        if modo == "paginas":
            # Mostrar campos de páginas
            self.paginas_frame_auto.pack(fill=tk.X, pady=(0, 20))
        else:  # ia
            # Esconder campos de páginas
            self.paginas_frame_auto.pack_forget()

    def _alternar_modo_manual(self):
        """Alterna entre sub-modos do modo manual (Completo vs Por Páginas)."""
        modo = self.modo_manual_tipo.get()

        if modo == "por_paginas":
            # Mostrar campos de páginas
            self.paginas_frame_manual.pack(fill=tk.X, pady=(0, 20))
        else:  # completo
            # Esconder campos de páginas
            self.paginas_frame_manual.pack_forget()

    def _atualizar_visual_submodo_auto(self):
        """Atualiza visual dos cards de sub-modo automático baseado na seleção."""
        modo = self.modo_automatico_tipo.get()

        # Resetar todos os cards para estado não selecionado
        self.card_ia.config(highlightthickness=2, highlightbackground=self.colors['border'])
        self.badge_ia.pack_forget()

        self.card_paginas_auto.config(highlightthickness=2, highlightbackground=self.colors['border'])
        self.badge_paginas_auto.pack_forget()

        # Destacar o card selecionado
        if modo == "ia":
            self.card_ia.config(highlightthickness=3, highlightbackground='#7C3AED')
            self.badge_ia.pack(pady=(15, 0))
        elif modo == "paginas":
            self.card_paginas_auto.config(highlightthickness=3, highlightbackground='#3B82F6')
            self.badge_paginas_auto.pack(pady=(15, 0))

    def _atualizar_visual_submodo_manual(self):
        """Atualiza visual dos cards de sub-modo manual baseado na seleção."""
        modo = self.modo_manual_tipo.get()

        # Resetar todos os cards para estado não selecionado
        self.card_completo.config(highlightthickness=2, highlightbackground=self.colors['border'])
        self.badge_completo.pack_forget()

        self.card_por_paginas_manual.config(highlightthickness=2, highlightbackground=self.colors['border'])
        self.badge_por_paginas_manual.pack_forget()

        # Destacar o card selecionado
        if modo == "completo":
            self.card_completo.config(highlightthickness=3, highlightbackground='#10B981')
            self.badge_completo.pack(pady=(15, 0))
        elif modo == "por_paginas":
            self.card_por_paginas_manual.config(highlightthickness=3, highlightbackground='#3B82F6')
            self.badge_por_paginas_manual.pack(pady=(15, 0))

    def _criar_modo_automatico_content(self):
        """Cria conteúdo do modo automático."""
        self.automatico_content = self._criar_card(self.content_frame)

        content = tk.Frame(self.automatico_content, bg=self.colors['bg_card'])
        content.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        # Descrição
        tk.Label(
            content,
            text="🚀  O sistema buscará automaticamente o arquivo na rede e processará tudo para você!",
            font=('Inter', 11),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 25))

        # ===== SELETOR DE SUB-MODO (CARDS CLICÁVEIS) =====
        submodo_container = tk.Frame(content, bg=self.colors['bg_card'])
        submodo_container.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            submodo_container,
            text="⚙️  Tipo de Processamento Automático",
            font=('Inter', 11, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(anchor=tk.W, pady=(0, 15))

        # Container para os cards lado a lado
        cards_frame = tk.Frame(submodo_container, bg=self.colors['bg_card'])
        cards_frame.pack(fill=tk.X)

        # ===== CARD 1: Por IA =====
        self.card_ia = tk.Frame(
            cards_frame,
            bg='#F3E8FF',
            highlightthickness=2,
            highlightbackground=self.colors['border'],
            cursor='hand2'
        )
        self.card_ia.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(0, 10))

        ia_content = tk.Frame(self.card_ia, bg='#F3E8FF')
        ia_content.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Ícone e título
        header_ia = tk.Frame(ia_content, bg='#F3E8FF')
        header_ia.pack(fill=tk.X)

        tk.Label(
            header_ia,
            text="🤖",
            font=('Segoe UI Emoji', 28),
            bg='#F3E8FF'
        ).pack(side=tk.LEFT, padx=(0, 8))

        title_frame_ia = tk.Frame(header_ia, bg='#F3E8FF')
        title_frame_ia.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(
            title_frame_ia,
            text="Por IA",
            font=('Inter', 12, 'bold'),
            fg='#7C3AED',
            bg='#F3E8FF',
            anchor=tk.W
        ).pack(anchor=tk.W)

        tk.Label(
            title_frame_ia,
            text="Automático",
            font=('Inter', 8),
            fg='#7C3AED',
            bg='#F3E8FF',
            anchor=tk.W
        ).pack(anchor=tk.W)

        # Descrição
        tk.Label(
            ia_content,
            text="IA detecta\npáginas",
            font=('Inter', 8),
            fg='#6B21A8',
            bg='#F3E8FF',
            justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(8, 0))

        # Badge de selecionado
        self.badge_ia = tk.Label(
            ia_content,
            text="✓ SELECIONADO",
            font=('Inter', 7, 'bold'),
            fg='white',
            bg='#7C3AED',
            padx=8,
            pady=3
        )

        # ===== CARD 2: Por Páginas =====
        self.card_paginas_auto = tk.Frame(
            cards_frame,
            bg='#DBEAFE',
            highlightthickness=3,
            highlightbackground='#3B82F6',
            cursor='hand2'
        )
        self.card_paginas_auto.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        paginas_auto_content = tk.Frame(self.card_paginas_auto, bg='#DBEAFE')
        paginas_auto_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Ícone e título
        header_paginas_auto = tk.Frame(paginas_auto_content, bg='#DBEAFE')
        header_paginas_auto.pack(fill=tk.X)

        tk.Label(
            header_paginas_auto,
            text="📄",
            font=('Segoe UI Emoji', 32),
            bg='#DBEAFE'
        ).pack(side=tk.LEFT, padx=(0, 10))

        title_frame_paginas_auto = tk.Frame(header_paginas_auto, bg='#DBEAFE')
        title_frame_paginas_auto.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(
            title_frame_paginas_auto,
            text="Por Páginas",
            font=('Inter', 13, 'bold'),
            fg='#1E40AF',
            bg='#DBEAFE',
            anchor=tk.W
        ).pack(anchor=tk.W)

        tk.Label(
            title_frame_paginas_auto,
            text="Manual",
            font=('Inter', 9),
            fg='#1E40AF',
            bg='#DBEAFE',
            anchor=tk.W
        ).pack(anchor=tk.W)

        # Descrição
        tk.Label(
            paginas_auto_content,
            text="Você especifica quais\npáginas extrair",
            font=('Inter', 9),
            fg='#1E3A8A',
            bg='#DBEAFE',
            justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(10, 0))

        # Badge de selecionado
        self.badge_paginas_auto = tk.Label(
            paginas_auto_content,
            text="✓ SELECIONADO",
            font=('Inter', 8, 'bold'),
            fg='white',
            bg='#3B82F6',
            padx=10,
            pady=4
        )

        # ===== EVENTOS DE CLIQUE =====
        def selecionar_ia(event=None):
            self.modo_automatico_tipo.set("ia")
            self._alternar_modo_automatico()
            self._atualizar_visual_submodo_auto()

        def selecionar_paginas(event=None):
            self.modo_automatico_tipo.set("paginas")
            self._alternar_modo_automatico()
            self._atualizar_visual_submodo_auto()

        # Bind card IA
        self.card_ia.bind('<Button-1>', selecionar_ia)
        for widget in ia_content.winfo_children():
            widget.bind('<Button-1>', selecionar_ia)
            for child in widget.winfo_children():
                child.bind('<Button-1>', selecionar_ia)

        # Bind card Páginas
        self.card_paginas_auto.bind('<Button-1>', selecionar_paginas)
        for widget in paginas_auto_content.winfo_children():
            widget.bind('<Button-1>', selecionar_paginas)
            for child in widget.winfo_children():
                child.bind('<Button-1>', selecionar_paginas)

        # Atualizar visual inicial
        self._atualizar_visual_submodo_auto()

        # ===== CAMPOS DE PÁGINAS (visível apenas no modo "paginas") =====
        self.paginas_frame_auto = tk.Frame(content, bg=self.colors['bg_card'])
        self.paginas_frame_auto.pack(fill=tk.X, pady=(0, 20))

        paginas_card = tk.Frame(
            self.paginas_frame_auto,
            bg='#FEF3C7',
            highlightthickness=2,
            highlightbackground='#FCD34D'
        )
        paginas_card.pack(fill=tk.X)

        paginas_card_content = tk.Frame(paginas_card, bg='#FEF3C7')
        paginas_card_content.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(
            paginas_card_content,
            text="📋  Especifique as Páginas para Extração",
            font=('Inter', 11, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 12))

        # Input páginas INCRA
        incra_pag_frame = tk.Frame(paginas_card_content, bg='#FEF3C7')
        incra_pag_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(
            incra_pag_frame,
            text="📄  Páginas do Memorial INCRA:",
            font=('Inter', 10, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 5))

        tk.Entry(
            incra_pag_frame,
            textvariable=self.paginas_incra_auto,
            font=('Inter', 10),
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(fill=tk.X, ipady=8, ipadx=10)

        tk.Label(
            incra_pag_frame,
            text="Ex: 1,2,4,7 (separe os números por vírgula)",
            font=('Inter', 8),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(3, 0))

        # Input páginas Projeto
        projeto_pag_frame = tk.Frame(paginas_card_content, bg='#FEF3C7')
        projeto_pag_frame.pack(fill=tk.X)

        tk.Label(
            projeto_pag_frame,
            text="📐  Páginas da Planta/Projeto:",
            font=('Inter', 10, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 5))

        tk.Entry(
            projeto_pag_frame,
            textvariable=self.paginas_projeto_auto,
            font=('Inter', 10),
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(fill=tk.X, ipady=8, ipadx=10)

        tk.Label(
            projeto_pag_frame,
            text="Ex: 5,6 (separe os números por vírgula)",
            font=('Inter', 8),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(3, 0))

        # ===== OPÇÃO DE SELEÇÃO MANUAL DE ARQUIVOS =====
        selecao_manual_card = tk.Frame(
            self.paginas_frame_auto,
            bg='#E0F2FE',
            highlightthickness=2,
            highlightbackground='#38BDF8'
        )
        selecao_manual_card.pack(fill=tk.X, pady=(15, 0))

        selecao_manual_content = tk.Frame(selecao_manual_card, bg='#E0F2FE')
        selecao_manual_content.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(
            selecao_manual_content,
            text="💡  Ou Selecione Arquivos Manualmente (Opcional)",
            font=('Inter', 10, 'bold'),
            fg='#0369A1',
            bg='#E0F2FE'
        ).pack(anchor=tk.W, pady=(0, 10))

        tk.Label(
            selecao_manual_content,
            text="Use esta opção se preferir selecionar o arquivo diretamente em vez de buscar pela prenotação",
            font=('Inter', 8),
            fg='#075985',
            bg='#E0F2FE'
        ).pack(anchor=tk.W, pady=(0, 10))

        # Seleção arquivo INCRA
        incra_manual_frame = tk.Frame(selecao_manual_content, bg='#E0F2FE')
        incra_manual_frame.pack(fill=tk.X, pady=(0, 8))

        tk.Label(
            incra_manual_frame,
            text="📄  Arquivo INCRA:",
            font=('Inter', 9, 'bold'),
            fg='#0369A1',
            bg='#E0F2FE'
        ).pack(anchor=tk.W, pady=(0, 5))

        incra_select_frame = tk.Frame(incra_manual_frame, bg='#E0F2FE')
        incra_select_frame.pack(fill=tk.X)

        self.incra_manual_auto_entry = tk.Entry(
            incra_select_frame,
            textvariable=self.arquivo_manual_incra_auto,
            font=('Inter', 9),
            state='readonly',
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=1,
            highlightthickness=0
        )
        self.incra_manual_auto_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, ipadx=8)

        tk.Button(
            incra_select_frame,
            text="📁 Selecionar",
            command=lambda: self._selecionar_arquivo_hibrido("incra"),
            font=('Inter', 9, 'bold'),
            bg='#0284C7',
            fg='white',
            relief=tk.FLAT,
            padx=12,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Seleção arquivo Projeto
        projeto_manual_frame = tk.Frame(selecao_manual_content, bg='#E0F2FE')
        projeto_manual_frame.pack(fill=tk.X)

        tk.Label(
            projeto_manual_frame,
            text="📐  Arquivo Projeto/Planta:",
            font=('Inter', 9, 'bold'),
            fg='#0369A1',
            bg='#E0F2FE'
        ).pack(anchor=tk.W, pady=(0, 5))

        projeto_select_frame = tk.Frame(projeto_manual_frame, bg='#E0F2FE')
        projeto_select_frame.pack(fill=tk.X)

        self.projeto_manual_auto_entry = tk.Entry(
            projeto_select_frame,
            textvariable=self.arquivo_manual_projeto_auto,
            font=('Inter', 9),
            state='readonly',
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=1,
            highlightthickness=0
        )
        self.projeto_manual_auto_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, ipadx=8)

        tk.Button(
            projeto_select_frame,
            text="📁 Selecionar",
            command=lambda: self._selecionar_arquivo_hibrido("projeto"),
            font=('Inter', 9, 'bold'),
            bg='#0284C7',
            fg='white',
            relief=tk.FLAT,
            padx=12,
            pady=6,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=(8, 0))

        # Botão grande de iniciar
        self.btn_iniciar_automatico = tk.Button(
            content,
            text="🚀  INICIAR BUSCA AUTOMÁTICA",
            command=self._iniciar_modo_automatico,
            font=('Inter', 14, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=40,
            pady=20,
            cursor='hand2',
            activebackground=self.colors['primary_dark'],
            activeforeground='white'
        )
        self.btn_iniciar_automatico.pack(pady=20)

        # Frame de preview (inicialmente oculto)
        self.preview_frame = tk.Frame(content, bg=self.colors['bg_card'])

        preview_title = tk.Label(
            self.preview_frame,
            text="👁️  Prévia dos Documentos Extraídos",
            font=('Inter', 12, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['bg_card']
        )
        preview_title.pack(pady=(20, 15))

        # Container para previews lado a lado
        preview_container = tk.Frame(self.preview_frame, bg=self.colors['bg_card'])
        preview_container.pack(fill=tk.BOTH, expand=True)

        # Preview INCRA
        incra_frame = tk.Frame(preview_container, bg=self.colors['bg_card'])
        incra_frame.pack(side=tk.LEFT, padx=15, expand=True)

        tk.Label(
            incra_frame,
            text="📄 Memorial INCRA",
            font=('Inter', 11, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 10))

        self.incra_preview_label = tk.Label(
            incra_frame,
            bg=self.colors['bg_light'],
            relief=tk.FLAT,
            highlightthickness=2,
            highlightbackground=self.colors['border']
        )
        self.incra_preview_label.pack()

        # Preview Projeto
        projeto_frame = tk.Frame(preview_container, bg=self.colors['bg_card'])
        projeto_frame.pack(side=tk.LEFT, padx=15, expand=True)

        tk.Label(
            projeto_frame,
            text="📐 Planta/Projeto",
            font=('Inter', 11, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 10))

        self.projeto_preview_label = tk.Label(
            projeto_frame,
            bg=self.colors['bg_light'],
            relief=tk.FLAT,
            highlightthickness=2,
            highlightbackground=self.colors['border']
        )
        self.projeto_preview_label.pack()

        # Botões de confirmação
        confirm_frame = tk.Frame(self.preview_frame, bg=self.colors['bg_card'])
        confirm_frame.pack(pady=25)

        tk.Button(
            confirm_frame,
            text="✅  CONTINUAR",
            command=self._confirmar_documentos_automaticos,
            font=('Inter', 12, 'bold'),
            bg=self.colors['success'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=12,
            cursor='hand2',
            activebackground='#059669'
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            confirm_frame,
            text="✋  FAZER MANUAL",
            command=self._alternar_para_manual,
            font=('Inter', 12, 'bold'),
            bg=self.colors['warning'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=12,
            cursor='hand2',
            activebackground='#D97706'
        ).pack(side=tk.LEFT, padx=10)

    def _criar_modo_manual_content(self):
        """Cria conteúdo do modo manual."""
        self.manual_content = self._criar_card(self.content_frame)

        content = tk.Frame(self.manual_content, bg=self.colors['bg_card'])
        content.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        # Descrição
        tk.Label(
            content,
            text="📁  Selecione manualmente os arquivos PDF para comparação",
            font=('Inter', 11),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 25))

        # ===== SELETOR DE SUB-MODO MANUAL (CARDS CLICÁVEIS) =====
        submodo_manual_container = tk.Frame(content, bg=self.colors['bg_card'])
        submodo_manual_container.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            submodo_manual_container,
            text="⚙️  Tipo de Processamento Manual",
            font=('Inter', 11, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(anchor=tk.W, pady=(0, 15))

        # Container para os cards lado a lado
        cards_manual_frame = tk.Frame(submodo_manual_container, bg=self.colors['bg_card'])
        cards_manual_frame.pack(fill=tk.X)

        # ===== CARD 1: Completo =====
        self.card_completo = tk.Frame(
            cards_manual_frame,
            bg='#D1FAE5',
            highlightthickness=3,
            highlightbackground='#10B981',
            cursor='hand2'
        )
        self.card_completo.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(0, 10))

        completo_content = tk.Frame(self.card_completo, bg='#D1FAE5')
        completo_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Ícone e título
        header_completo = tk.Frame(completo_content, bg='#D1FAE5')
        header_completo.pack(fill=tk.X)

        tk.Label(
            header_completo,
            text="📄",
            font=('Segoe UI Emoji', 32),
            bg='#D1FAE5'
        ).pack(side=tk.LEFT, padx=(0, 10))

        title_frame_completo = tk.Frame(header_completo, bg='#D1FAE5')
        title_frame_completo.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(
            title_frame_completo,
            text="Completo",
            font=('Inter', 13, 'bold'),
            fg='#059669',
            bg='#D1FAE5',
            anchor=tk.W
        ).pack(anchor=tk.W)

        tk.Label(
            title_frame_completo,
            text="Padrão",
            font=('Inter', 9),
            fg='#059669',
            bg='#D1FAE5',
            anchor=tk.W
        ).pack(anchor=tk.W)

        # Descrição
        tk.Label(
            completo_content,
            text="Processa arquivo\ninteiro",
            font=('Inter', 9),
            fg='#047857',
            bg='#D1FAE5',
            justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(10, 0))

        # Badge de selecionado
        self.badge_completo = tk.Label(
            completo_content,
            text="✓ SELECIONADO",
            font=('Inter', 8, 'bold'),
            fg='white',
            bg='#10B981',
            padx=10,
            pady=4
        )

        # ===== CARD 2: Por Páginas =====
        self.card_por_paginas_manual = tk.Frame(
            cards_manual_frame,
            bg='#DBEAFE',
            highlightthickness=2,
            highlightbackground=self.colors['border'],
            cursor='hand2'
        )
        self.card_por_paginas_manual.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        por_paginas_manual_content = tk.Frame(self.card_por_paginas_manual, bg='#DBEAFE')
        por_paginas_manual_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Ícone e título
        header_por_paginas_manual = tk.Frame(por_paginas_manual_content, bg='#DBEAFE')
        header_por_paginas_manual.pack(fill=tk.X)

        tk.Label(
            header_por_paginas_manual,
            text="📋",
            font=('Segoe UI Emoji', 32),
            bg='#DBEAFE'
        ).pack(side=tk.LEFT, padx=(0, 10))

        title_frame_por_paginas_manual = tk.Frame(header_por_paginas_manual, bg='#DBEAFE')
        title_frame_por_paginas_manual.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(
            title_frame_por_paginas_manual,
            text="Por Páginas",
            font=('Inter', 13, 'bold'),
            fg='#1E40AF',
            bg='#DBEAFE',
            anchor=tk.W
        ).pack(anchor=tk.W)

        tk.Label(
            title_frame_por_paginas_manual,
            text="Seletivo",
            font=('Inter', 9),
            fg='#1E40AF',
            bg='#DBEAFE',
            anchor=tk.W
        ).pack(anchor=tk.W)

        # Descrição
        tk.Label(
            por_paginas_manual_content,
            text="Você especifica\npáginas",
            font=('Inter', 9),
            fg='#1E3A8A',
            bg='#DBEAFE',
            justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(10, 0))

        # Badge de selecionado
        self.badge_por_paginas_manual = tk.Label(
            por_paginas_manual_content,
            text="✓ SELECIONADO",
            font=('Inter', 8, 'bold'),
            fg='white',
            bg='#3B82F6',
            padx=10,
            pady=4
        )

        # ===== EVENTOS DE CLIQUE =====
        def selecionar_completo(event=None):
            self.modo_manual_tipo.set("completo")
            self._alternar_modo_manual()
            self._atualizar_visual_submodo_manual()

        def selecionar_por_paginas_manual(event=None):
            self.modo_manual_tipo.set("por_paginas")
            self._alternar_modo_manual()
            self._atualizar_visual_submodo_manual()

        # Bind card Completo
        self.card_completo.bind('<Button-1>', selecionar_completo)
        for widget in completo_content.winfo_children():
            widget.bind('<Button-1>', selecionar_completo)
            for child in widget.winfo_children():
                child.bind('<Button-1>', selecionar_completo)

        # Bind card Por Páginas Manual
        self.card_por_paginas_manual.bind('<Button-1>', selecionar_por_paginas_manual)
        for widget in por_paginas_manual_content.winfo_children():
            widget.bind('<Button-1>', selecionar_por_paginas_manual)
            for child in widget.winfo_children():
                child.bind('<Button-1>', selecionar_por_paginas_manual)

        # Atualizar visual inicial
        self._atualizar_visual_submodo_manual()

        # ===== CAMPOS DE PÁGINAS MANUAL (visível apenas no modo "por_paginas") =====
        self.paginas_frame_manual = tk.Frame(content, bg=self.colors['bg_card'])

        paginas_manual_card = tk.Frame(
            self.paginas_frame_manual,
            bg='#FEF3C7',
            highlightthickness=2,
            highlightbackground='#FCD34D'
        )
        paginas_manual_card.pack(fill=tk.X)

        paginas_manual_card_content = tk.Frame(paginas_manual_card, bg='#FEF3C7')
        paginas_manual_card_content.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(
            paginas_manual_card_content,
            text="📋  Especifique as Páginas para Extração",
            font=('Inter', 11, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 12))

        # Input páginas INCRA
        incra_pag_manual_frame = tk.Frame(paginas_manual_card_content, bg='#FEF3C7')
        incra_pag_manual_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(
            incra_pag_manual_frame,
            text="📄  Páginas do Memorial INCRA:",
            font=('Inter', 10, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 5))

        tk.Entry(
            incra_pag_manual_frame,
            textvariable=self.paginas_incra_manual,
            font=('Inter', 10),
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(fill=tk.X, ipady=8, ipadx=10)

        tk.Label(
            incra_pag_manual_frame,
            text="Ex: 1,2,4,7 (separe os números por vírgula)",
            font=('Inter', 8),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(3, 0))

        # Input páginas Projeto
        projeto_pag_manual_frame = tk.Frame(paginas_manual_card_content, bg='#FEF3C7')
        projeto_pag_manual_frame.pack(fill=tk.X)

        tk.Label(
            projeto_pag_manual_frame,
            text="📐  Páginas da Planta/Projeto:",
            font=('Inter', 10, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 5))

        tk.Entry(
            projeto_pag_manual_frame,
            textvariable=self.paginas_projeto_manual,
            font=('Inter', 10),
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(fill=tk.X, ipady=8, ipadx=10)

        tk.Label(
            projeto_pag_manual_frame,
            text="Ex: 5,6 (separe os números por vírgula)",
            font=('Inter', 8),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(3, 0))

        # Seleção INCRA
        incra_card = tk.Frame(
            content,
            bg='#FEF3C7',
            highlightthickness=2,
            highlightbackground='#FCD34D'
        )
        incra_card.pack(fill=tk.X, pady=10)

        incra_content = tk.Frame(incra_card, bg='#FEF3C7')
        incra_content.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(
            incra_content,
            text="📄  Memorial INCRA",
            font=('Inter', 12, 'bold'),
            fg='#92400E',
            bg='#FEF3C7'
        ).pack(anchor=tk.W, pady=(0, 10))

        incra_input_frame = tk.Frame(incra_content, bg='#FEF3C7')
        incra_input_frame.pack(fill=tk.X)

        tk.Entry(
            incra_input_frame,
            textvariable=self.incra_path,
            font=('Inter', 10),
            state='readonly',
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, ipadx=10)

        tk.Button(
            incra_input_frame,
            text="📁 Selecionar",
            command=lambda: self._selecionar_arquivo(self.incra_path, "INCRA"),
            font=('Inter', 10, 'bold'),
            bg='#F59E0B',
            fg='white',
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=(10, 0))

        # Botão Arquivos Separados para INCRA
        tk.Button(
            incra_input_frame,
            text="📚 Arquivos Separados",
            command=lambda: self._selecionar_multiplos_arquivos(self.incra_path, "INCRA"),
            font=('Inter', 9, 'bold'),
            bg='#D97706',
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=(10, 0))

        # Seleção Projeto
        projeto_card = tk.Frame(
            content,
            bg='#DBEAFE',
            highlightthickness=2,
            highlightbackground='#60A5FA'
        )
        projeto_card.pack(fill=tk.X, pady=10)

        projeto_content = tk.Frame(projeto_card, bg='#DBEAFE')
        projeto_content.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(
            projeto_content,
            text="📐  Planta/Projeto",
            font=('Inter', 12, 'bold'),
            fg='#1E40AF',
            bg='#DBEAFE'
        ).pack(anchor=tk.W, pady=(0, 10))

        projeto_input_frame = tk.Frame(projeto_content, bg='#DBEAFE')
        projeto_input_frame.pack(fill=tk.X)

        tk.Entry(
            projeto_input_frame,
            textvariable=self.projeto_path,
            font=('Inter', 10),
            state='readonly',
            relief=tk.SOLID,
            bg='white',
            fg=self.colors['text_dark'],
            borderwidth=2,
            highlightthickness=0
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, ipadx=10)

        tk.Button(
            projeto_input_frame,
            text="📁 Selecionar",
            command=lambda: self._selecionar_arquivo(self.projeto_path, "Projeto"),
            font=('Inter', 10, 'bold'),
            bg='#3B82F6',
            fg='white',
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=(10, 0))

        # Botão Arquivos Separados para Projeto
        tk.Button(
            projeto_input_frame,
            text="📚 Arquivos Separados",
            command=lambda: self._selecionar_multiplos_arquivos(self.projeto_path, "Projeto"),
            font=('Inter', 9, 'bold'),
            bg='#2563EB',
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=(10, 0))

        # Botão de comparação
        tk.Button(
            content,
            text="🔍  COMPARAR DOCUMENTOS",
            command=self._comparar_manual,
            font=('Inter', 14, 'bold'),
            bg=self.colors['secondary'],
            fg='white',
            relief=tk.FLAT,
            padx=40,
            pady=20,
            cursor='hand2',
            activebackground='#DB2777'
        ).pack(pady=30)

    def _validar_numero(self, valor):
        """Valida entrada para aceitar apenas números."""
        return valor == "" or valor.isdigit()

    def _carregar_api_key(self):
        """Carrega API key salva e atualiza interface."""
        api_key = self.config_manager.get_api_key()
        if api_key:
            self.api_status_label.config(
                text="✅ Configurada",
                fg=self.colors['success']
            )
        else:
            self.api_status_label.config(
                text="⭕ Não configurada",
                fg=self.colors['danger']
            )

    def _criar_janela_progresso(self, titulo="Processando..."):
        """Cria janela de progresso com barra e informações."""
        if self.progress_window is not None:
            try:
                self.progress_window.destroy()
            except:
                pass

        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title(titulo)
        self.progress_window.geometry("600x300")
        self.progress_window.configure(bg=self.colors['bg_card'])
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        self.progress_window.resizable(False, False)

        # Centralizar na tela
        self.progress_window.update_idletasks()
        x = (self.progress_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (self.progress_window.winfo_screenheight() // 2) - (300 // 2)
        self.progress_window.geometry(f"600x300+{x}+{y}")

        # Desabilitar fechar a janela
        self.progress_window.protocol("WM_DELETE_WINDOW", lambda: None)

        # Frame principal
        main_frame = tk.Frame(self.progress_window, bg=self.colors['bg_card'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)

        # Ícone e título
        icon_label = tk.Label(
            main_frame,
            text="⏳",
            font=('Segoe UI Emoji', 48),
            bg=self.colors['bg_card']
        )
        icon_label.pack(pady=(0, 20))

        # Label de status principal
        self.progress_label = tk.Label(
            main_frame,
            text="Iniciando processamento...",
            font=('Inter', 14, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['bg_card']
        )
        self.progress_label.pack(pady=(0, 10))

        # Label de detalhes
        self.progress_detail_label = tk.Label(
            main_frame,
            text="",
            font=('Inter', 10),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card'],
            wraplength=500
        )
        self.progress_detail_label.pack(pady=(0, 25))

        # Frame para barra de progresso
        progress_frame = tk.Frame(main_frame, bg=self.colors['bg_card'])
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        # Barra de progresso
        style = ttk.Style()
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor='#E5E7EB',
            background=self.colors['primary'],
            bordercolor=self.colors['border'],
            lightcolor=self.colors['primary'],
            darkcolor=self.colors['primary_dark']
        )

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            style="Custom.Horizontal.TProgressbar",
            orient='horizontal',
            length=500,
            mode='determinate',
            maximum=100
        )
        self.progress_bar.pack()

        # Porcentagem
        self.progress_percent_label = tk.Label(
            main_frame,
            text="0%",
            font=('Inter', 11, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['bg_card']
        )
        self.progress_percent_label.pack()

        # Forçar atualização
        self.progress_window.update()

    def _atualizar_progresso(self, porcentagem: int, texto: str, detalhe: str = ""):
        """Atualiza a janela de progresso."""
        if self.progress_window and self.progress_window.winfo_exists():
            try:
                self.progress_bar['value'] = porcentagem
                self.progress_label.config(text=texto)
                self.progress_detail_label.config(text=detalhe)
                self.progress_percent_label.config(text=f"{porcentagem}%")
                self.progress_window.update()
            except:
                pass

    def _fechar_janela_progresso(self):
        """Fecha a janela de progresso."""
        if self.progress_window:
            try:
                self.progress_window.grab_release()
                self.progress_window.destroy()
                self.progress_window = None
            except:
                pass

    def _abrir_config_api(self):
        """Abre janela para configurar API key."""
        config_window = tk.Toplevel(self.root)
        config_window.title("⚙️ Configuração da API Key")
        config_window.geometry("750x550")
        config_window.configure(bg=self.colors['bg_card'])
        config_window.transient(self.root)
        config_window.grab_set()

        main_frame = tk.Frame(config_window, bg=self.colors['bg_card'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        # Título
        tk.Label(
            main_frame,
            text="🔑  Configuração da API Key do Gemini",
            font=('Inter', 16, 'bold'),
            fg=self.colors['primary'],
            bg=self.colors['bg_card']
        ).pack(pady=(0, 10))

        tk.Label(
            main_frame,
            text="Insira sua API key do Google Gemini abaixo.\nEla será salva de forma segura em seu computador.",
            font=('Inter', 10),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card'],
            justify=tk.CENTER
        ).pack(pady=10)

        # Campo de entrada com borda destacada
        api_var = tk.StringVar(value=self.config_manager.get_api_key())

        entry_frame = tk.Frame(
            main_frame,
            bg=self.colors['bg_card'],
            highlightbackground=self.colors['primary'],
            highlightthickness=2,
            relief=tk.SOLID
        )
        entry_frame.pack(fill=tk.X, pady=20, padx=10)

        entry_content = tk.Frame(entry_frame, bg=self.colors['bg_card'])
        entry_content.pack(fill=tk.X, padx=15, pady=15)

        tk.Label(
            entry_content,
            text="API Key:",
            font=('Inter', 12, 'bold'),
            fg=self.colors['text_dark'],
            bg=self.colors['bg_card']
        ).pack(anchor=tk.W, pady=(0, 10))

        api_entry = tk.Entry(
            entry_content,
            textvariable=api_var,
            font=('Inter', 11),
            show="●",
            relief=tk.SOLID,
            bg='#F3F4F6',
            fg=self.colors['text_dark'],
            insertbackground=self.colors['primary'],
            borderwidth=2,
            highlightthickness=0
        )
        api_entry.pack(fill=tk.X, ipady=12, ipadx=10)
        api_entry.focus()

        # Instrução adicional
        tk.Label(
            main_frame,
            text="💡 Depois de inserir a chave, clique em 'SALVAR CONFIGURAÇÃO' abaixo",
            font=('Inter', 9, 'italic'),
            fg=self.colors['info'],
            bg=self.colors['bg_card']
        ).pack(pady=(5, 15))

        # Botões grandes e destacados
        btn_frame = tk.Frame(main_frame, bg=self.colors['bg_card'])
        btn_frame.pack(pady=20)

        def salvar_api():
            key = api_var.get().strip()
            if key:
                self.config_manager.set_api_key(key)
                self._carregar_api_key()
                messagebox.showinfo("✅ Sucesso", "API Key salva com sucesso!")
                config_window.destroy()
            else:
                messagebox.showwarning("⚠️ Aviso", "Por favor, insira uma API Key válida.")

        # Botão de salvar grande e destacado
        tk.Button(
            btn_frame,
            text="💾  SALVAR CONFIGURAÇÃO",
            command=salvar_api,
            font=('Inter', 13, 'bold'),
            bg=self.colors['success'],
            fg='white',
            relief=tk.FLAT,
            padx=40,
            pady=15,
            cursor='hand2',
            activebackground='#059669',
            activeforeground='white'
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="❌  Cancelar",
            command=config_window.destroy,
            font=('Inter', 11, 'bold'),
            bg=self.colors['text_medium'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=15,
            cursor='hand2',
            activebackground='#4B5563',
            activeforeground='white'
        ).pack(side=tk.LEFT, padx=10)

        # Link para obter API key
        link_frame = tk.Frame(main_frame, bg=self.colors['bg_card'])
        link_frame.pack(pady=(20, 0))

        tk.Label(
            link_frame,
            text="❓ Não tem uma API Key?",
            font=('Inter', 9),
            fg=self.colors['text_medium'],
            bg=self.colors['bg_card']
        ).pack(side=tk.LEFT, padx=(0, 5))

        link_label = tk.Label(
            link_frame,
            text="Clique aqui para obter",
            font=('Inter', 9, 'bold', 'underline'),
            fg=self.colors['info'],
            bg=self.colors['bg_card'],
            cursor='hand2'
        )
        link_label.pack(side=tk.LEFT)
        link_label.bind('<Button-1>', lambda e: webbrowser.open('https://makersuite.google.com/app/apikey'))

    def _selecionar_arquivo(self, variavel, tipo):
        """Abre diálogo para selecionar arquivo PDF."""
        filename = filedialog.askopenfilename(
            title=f"Selecionar arquivo {tipo}",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            variavel.set(filename)

    def _selecionar_arquivo_hibrido(self, tipo):
        """Abre diálogo para selecionar arquivo PDF manualmente no modo automático (híbrido)."""
        nome_tipo = "Memorial INCRA" if tipo == "incra" else "Planta/Projeto"
        filename = filedialog.askopenfilename(
            title=f"Selecionar arquivo {nome_tipo} manualmente",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            if tipo == "incra":
                self.arquivo_manual_incra_auto.set(filename)
                messagebox.showinfo(
                    "Arquivo Selecionado",
                    f"✅ Arquivo INCRA selecionado:\n\n{Path(filename).name}\n\n"
                    "O sistema usará este arquivo em vez de buscar pela prenotação."
                )
            else:  # projeto
                self.arquivo_manual_projeto_auto.set(filename)
                messagebox.showinfo(
                    "Arquivo Selecionado",
                    f"✅ Arquivo Projeto selecionado:\n\n{Path(filename).name}\n\n"
                    "O sistema usará este arquivo em vez de buscar pela prenotação."
                )

    def _selecionar_multiplos_arquivos(self, variavel, tipo):
        """Abre diálogo para selecionar múltiplos arquivos PDF e faz o merge."""
        filenames = filedialog.askopenfilenames(
            title=f"Selecionar múltiplos arquivos {tipo} (serão mesclados)",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )

        if filenames and len(filenames) > 0:
            if len(filenames) == 1:
                # Se selecionou apenas um arquivo, não precisa fazer merge
                variavel.set(filenames[0])
                messagebox.showinfo("Arquivo Selecionado", f"1 arquivo selecionado para {tipo}.")
            else:
                # Fazer merge dos PDFs
                try:
                    output_path = self._merge_pdfs(list(filenames), tipo)
                    variavel.set(output_path)
                    messagebox.showinfo(
                        "Arquivos Mesclados",
                        f"{len(filenames)} arquivos foram mesclados com sucesso!\n\nArquivo final: {Path(output_path).name}"
                    )
                except Exception as e:
                    messagebox.showerror("Erro ao Mesclar", f"Erro ao mesclar arquivos:\n{e}")

    def _merge_pdfs(self, pdf_files: List[str], tipo: str) -> str:
        """Mescla múltiplos arquivos PDF em um único arquivo."""
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Nome do arquivo de saída
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{tipo.lower()}_mesclado_{timestamp}.pdf"
        output_path = output_dir / output_filename

        # Criar o merger
        pdf_merger = PyPDF2.PdfMerger()

        try:
            # Adicionar cada PDF ao merger
            for pdf_file in pdf_files:
                pdf_merger.append(pdf_file)

            # Escrever o arquivo mesclado
            with open(output_path, 'wb') as output_file:
                pdf_merger.write(output_file)

            pdf_merger.close()

            print(f"✅ {len(pdf_files)} arquivos mesclados em: {output_path}")
            return str(output_path)

        except Exception as e:
            pdf_merger.close()
            raise e

    def _selecionar_arquivo_manual(self):
        """Abre diálogo para selecionar arquivo PDF manualmente (para modo paginas_manual)."""
        filename = filedialog.askopenfilename(
            title="Selecionar arquivo PDF completo (Memorial + Planta)",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            self.arquivo_manual_path.set(filename)
            messagebox.showinfo(
                "Arquivo Selecionado",
                f"Arquivo selecionado:\n\n{Path(filename).name}\n\nAgora especifique as páginas do Memorial e da Planta."
            )

    def _atualizar_status(self, mensagem: str):
        """Atualiza a barra de status."""
        # Detectar tipo de mensagem e ajustar cor
        if "✅" in mensagem or "sucesso" in mensagem.lower():
            cor = self.colors['success']
        elif "❌" in mensagem or "erro" in mensagem.lower():
            cor = self.colors['danger']
        elif "🔄" in mensagem or "processando" in mensagem.lower():
            cor = self.colors['info']
        else:
            cor = self.colors['text_dark']

        self.status_label.config(text=mensagem, fg=cor)
        self.root.update_idletasks()

    def _desabilitar_botoes(self):
        """Desabilita botões durante o processamento."""
        self.btn_iniciar_automatico.config(state='disabled', bg=self.colors['text_light'])

    def _habilitar_botoes(self):
        """Reabilita botões após o processamento."""
        self.btn_iniciar_automatico.config(state='normal', bg=self.colors['primary'])

    # ========== MODO MANUAL ==========

    def _comparar_manual(self):
        """Executa comparação no modo manual."""
        if not self._validar_entrada_manual():
            return

        def executar():
            try:
                self._desabilitar_botoes()

                # Criar janela de progresso
                self._criar_janela_progresso("Comparação em Andamento")

                modo = self.modo_manual_tipo.get()
                incra_file = self.incra_path.get()
                projeto_file = self.projeto_path.get()

                # Se for modo "por_paginas", extrair páginas especificadas primeiro
                if modo == "por_paginas":
                    # Etapa 1: Extrair páginas do INCRA (0-20%)
                    self._atualizar_progresso(
                        5,
                        "Extraindo páginas do INCRA...",
                        f"Extraindo páginas: {self.paginas_incra_manual.get()}"
                    )
                    self._atualizar_status("📄 Extraindo páginas do Memorial INCRA...")
                    incra_file = self._extrair_paginas_manual(
                        self.incra_path.get(),
                        self.paginas_incra_manual.get(),
                        "incra_manual_extraido.pdf"
                    )
                    self._atualizar_progresso(15, "Páginas do INCRA extraídas!", "")

                    # Etapa 2: Extrair páginas do Projeto (20-35%)
                    self._atualizar_progresso(
                        20,
                        "Extraindo páginas do Projeto...",
                        f"Extraindo páginas: {self.paginas_projeto_manual.get()}"
                    )
                    self._atualizar_status("📐 Extraindo páginas da Planta/Projeto...")
                    projeto_file = self._extrair_paginas_manual(
                        self.projeto_path.get(),
                        self.paginas_projeto_manual.get(),
                        "projeto_manual_extraido.pdf"
                    )
                    self._atualizar_progresso(35, "Páginas do Projeto extraídas!", "")
                else:
                    # Modo completo: Preparação
                    self._atualizar_progresso(
                        5,
                        "Preparando para processar...",
                        "Verificando arquivos e configurações iniciais"
                    )
                    self._atualizar_status("🔄 Processando documentos...")

                # Etapa 3: Extrair dados INCRA (35-55%)
                self._atualizar_progresso(
                    40,
                    "Extraindo dados do INCRA...",
                    "Lendo PDF do INCRA e convertendo para Excel usando IA"
                )
                self._atualizar_status("📄 Extraindo dados do INCRA...")
                self.incra_excel_path, self.incra_data = self._extrair_pdf_para_excel(
                    incra_file, "incra"
                )
                self._atualizar_progresso(55, "INCRA extraído com sucesso!", "")

                # Etapa 4: Extrair dados Projeto (55-75%)
                self._atualizar_progresso(
                    60,
                    "Extraindo dados do Projeto...",
                    "Lendo PDF do Projeto/Planta e convertendo para Excel usando IA"
                )
                self._atualizar_status("📐 Extraindo dados do Projeto...")
                self.projeto_excel_path, self.projeto_data = self._extrair_pdf_para_excel(
                    projeto_file, "normal"
                )
                self._atualizar_progresso(75, "Projeto extraído com sucesso!", "")

                # Etapa 4: Gerar relatório (70-90%)
                self._atualizar_progresso(
                    75,
                    "Gerando relatório de comparação...",
                    "Comparando vértices e segmentos entre INCRA e Projeto"
                )
                self._atualizar_status("📊 Gerando relatório de comparação...")
                relatorio = self._construir_relatorio_comparacao(True, False)
                self._atualizar_progresso(85, "Relatório gerado!", "")

                # Etapa 5: Salvar e abrir (90-100%)
                self._atualizar_progresso(
                    90,
                    "Salvando relatório...",
                    "Salvando arquivo HTML e abrindo no navegador"
                )
                self._salvar_e_abrir_relatorio(relatorio)

                # Mostrar resumo
                self._atualizar_progresso(95, "Finalizando...", "Atualizando interface")
                self._mostrar_resumo_no_texto()

                # Concluído
                self._atualizar_progresso(100, "✅ Concluído!", "Comparação realizada com sucesso")
                self._atualizar_status("✅ Comparação concluída com sucesso!")

                # Esperar 1.5 segundos para o usuário ver a conclusão
                import time
                time.sleep(1.5)

            except Exception as e:
                self._atualizar_status(f"❌ Erro: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao processar documentos:\n\n{str(e)}")
            finally:
                self._fechar_janela_progresso()
                self._habilitar_botoes()

        # Executar em thread separada
        threading.Thread(target=executar, daemon=True).start()

    def _validar_entrada_manual(self) -> bool:
        """Valida entradas do modo manual."""
        api_key = self.config_manager.get_api_key()
        if not api_key:
            messagebox.showerror("Erro", "Por favor, configure a API Key primeiro.")
            return False

        if not self.incra_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo INCRA.")
            return False

        if not self.projeto_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo Projeto/Planta.")
            return False

        # Validação específica para modo "por_paginas"
        if self.modo_manual_tipo.get() == "por_paginas":
            if not self.paginas_incra_manual.get().strip():
                messagebox.showerror("Erro", "Por favor, especifique as páginas do Memorial INCRA.")
                return False

            if not self.paginas_projeto_manual.get().strip():
                messagebox.showerror("Erro", "Por favor, especifique as páginas da Planta/Projeto.")
                return False

        if not self.numero_prenotacao.get():
            messagebox.showerror("Erro", "Por favor, insira o Número de Prenotação.")
            return False

        return True

    # ========== MODO AUTOMÁTICO ==========

    def _iniciar_modo_automatico(self):
        """Inicia o processo automático."""
        if not self._validar_entrada_automatico():
            return

        def executar():
            try:
                self._desabilitar_botoes()

                # Criar janela de progresso
                self._criar_janela_progresso("Processamento em Andamento")

                modo = self.modo_automatico_tipo.get()

                # Verificar se há arquivos selecionados manualmente
                tem_incra_manual = bool(self.arquivo_manual_incra_auto.get())
                tem_projeto_manual = bool(self.arquivo_manual_projeto_auto.get())

                # Determinar arquivo fonte para INCRA
                if tem_incra_manual:
                    # INCRA: Arquivo selecionado manualmente
                    self._atualizar_progresso(
                        5,
                        "Usando arquivo INCRA selecionado manualmente...",
                        f"Arquivo: {Path(self.arquivo_manual_incra_auto.get()).name}"
                    )
                    self._atualizar_status("📄 Usando arquivo INCRA manual...")
                    pdf_incra_path = self.arquivo_manual_incra_auto.get()
                    self._atualizar_progresso(15, "Arquivo INCRA pronto!", "")
                else:
                    # INCRA: Buscar automaticamente pela prenotação
                    self._atualizar_progresso(
                        5,
                        "Buscando arquivo TIFF na rede...",
                        f"Procurando documento na rede (Prenotação: {self.numero_prenotacao.get()})"
                    )
                    self._atualizar_status("🔍 Buscando arquivo TIFF na rede...")
                    tiff_path = self._buscar_arquivo_tiff()

                    if not tiff_path:
                        raise Exception("Arquivo TIFF não encontrado na rede.")

                    self._atualizar_progresso(10, "Arquivo encontrado!", f"Localizado: {tiff_path}")

                    # Converter TIFF para PDF
                    self._atualizar_progresso(
                        12,
                        "Convertendo TIFF para PDF...",
                        "Copiando arquivo e convertendo formato"
                    )
                    self._atualizar_status("📋 Copiando e convertendo TIFF para PDF...")
                    pdf_incra_path = self._converter_tiff_para_pdf(tiff_path)
                    self._atualizar_progresso(15, "PDF criado com sucesso!", "")

                # Determinar arquivo fonte para Projeto
                if tem_projeto_manual:
                    # Projeto: Arquivo selecionado manualmente
                    self._atualizar_progresso(
                        18,
                        "Usando arquivo Projeto selecionado manualmente...",
                        f"Arquivo: {Path(self.arquivo_manual_projeto_auto.get()).name}"
                    )
                    self._atualizar_status("📐 Usando arquivo Projeto manual...")
                    pdf_projeto_path = self.arquivo_manual_projeto_auto.get()
                    self._atualizar_progresso(25, "Arquivo Projeto pronto!", "")
                else:
                    # Projeto: Usar mesmo PDF do INCRA
                    pdf_projeto_path = pdf_incra_path
                    self._atualizar_progresso(25, "Usando mesmo PDF para Projeto", "")

                # Etapa 3: Extrair Memorial INCRA (25-50%)
                if modo == "ia":
                    # Modo IA: Usar Gemini para detectar páginas
                    self._atualizar_progresso(
                        30,
                        "Extraindo Memorial INCRA...",
                        "Usando IA para identificar e extrair páginas do Memorial INCRA"
                    )
                    self._atualizar_status("📄 Extraindo Memorial INCRA com IA...")
                    self.pdf_extraido_incra = self._extrair_memorial_incra_do_pdf(pdf_incra_path)
                else:
                    # Modo Páginas: Extrair páginas especificadas pelo usuário
                    self._atualizar_progresso(
                        30,
                        "Extraindo Memorial INCRA...",
                        f"Extraindo páginas especificadas: {self.paginas_incra_auto.get()}"
                    )
                    self._atualizar_status("📄 Extraindo Memorial INCRA (páginas especificadas)...")
                    self.pdf_extraido_incra = self._extrair_paginas_manual(
                        pdf_incra_path,
                        self.paginas_incra_auto.get(),
                        "memorial_incra_extraido.pdf"
                    )

                if not self.pdf_extraido_incra:
                    raise Exception("Falha ao extrair Memorial INCRA")

                self._atualizar_progresso(50, "Memorial INCRA extraído!", "")

                # Etapa 4: Extrair Planta/Projeto (50-75%)
                if modo == "ia":
                    # Modo IA: Usar Gemini para detectar páginas
                    self._atualizar_progresso(
                        55,
                        "Extraindo Planta/Projeto...",
                        "Usando IA para identificar e extrair páginas da Planta/Projeto"
                    )
                    self._atualizar_status("📐 Extraindo Planta/Projeto com IA...")
                    self.pdf_extraido_projeto = self._extrair_projeto_do_pdf(pdf_projeto_path)
                else:
                    # Modo Páginas: Extrair páginas especificadas pelo usuário
                    self._atualizar_progresso(
                        55,
                        "Extraindo Planta/Projeto...",
                        f"Extraindo páginas especificadas: {self.paginas_projeto_auto.get()}"
                    )
                    self._atualizar_status("📐 Extraindo Planta/Projeto (páginas especificadas)...")
                    self.pdf_extraido_projeto = self._extrair_paginas_manual(
                        pdf_projeto_path,
                        self.paginas_projeto_auto.get(),
                        "projeto_extraido.pdf"
                    )

                if not self.pdf_extraido_projeto:
                    raise Exception("Falha ao extrair Planta/Projeto")

                self._atualizar_progresso(75, "Planta/Projeto extraída!", "")

                # Etapa 5: Salvar backups (75-85%)
                self._atualizar_progresso(
                    80,
                    "Salvando backups...",
                    "Criando cópias de segurança dos PDFs extraídos"
                )
                self._atualizar_status("💾 Salvando backups...")
                self._salvar_backups_pdfs()
                self._atualizar_progresso(85, "Backups salvos!", "")

                # Etapa 6: Gerar previews (85-100%)
                self._atualizar_progresso(
                    90,
                    "Gerando prévias...",
                    "Criando miniaturas dos documentos para visualização"
                )
                self._atualizar_status("👁️ Gerando prévias...")
                self._gerar_previews()
                self._atualizar_progresso(95, "Prévias geradas!", "")

                # Concluído
                self._atualizar_progresso(
                    100,
                    "✅ Documentos extraídos!",
                    "Verifique as prévias e confirme para continuar"
                )

                # Esperar 1 segundo para o usuário ver a conclusão
                import time
                time.sleep(1)

                # Fechar janela de progresso
                self._fechar_janela_progresso()

                # Mostrar frame de preview
                self.preview_frame.pack(fill=tk.BOTH, expand=True, pady=20)

                self._atualizar_status("✅ Documentos extraídos! Verifique as prévias.")

            except Exception as e:
                self._fechar_janela_progresso()
                self._atualizar_status(f"❌ Erro: {str(e)}")
                messagebox.showerror("Erro", f"Erro no modo automático:\n\n{str(e)}")
                self._habilitar_botoes()

        # Executar em thread separada
        threading.Thread(target=executar, daemon=True).start()

    def _validar_entrada_automatico(self) -> bool:
        """Valida entradas do modo automático (suporta modo híbrido)."""
        modo = self.modo_automatico_tipo.get()

        # Verificar se há arquivos selecionados manualmente
        tem_arquivo_incra_manual = bool(self.arquivo_manual_incra_auto.get())
        tem_arquivo_projeto_manual = bool(self.arquivo_manual_projeto_auto.get())

        # Validações específicas por sub-modo
        if modo == "ia":
            # Modo IA: precisa de prenotação e API Key
            if not self.numero_prenotacao.get():
                messagebox.showerror("Erro", "Por favor, insira o Número de Prenotação.")
                return False

            api_key = self.config_manager.get_api_key()
            if not api_key:
                messagebox.showerror("Erro", "Por favor, configure a API Key primeiro.")
                return False

        elif modo == "paginas":
            # Modo Páginas: precisa de números de páginas
            if not self.paginas_incra_auto.get().strip():
                messagebox.showerror("Erro", "Por favor, especifique as páginas do Memorial INCRA.")
                return False

            if not self.paginas_projeto_auto.get().strip():
                messagebox.showerror("Erro", "Por favor, especifique as páginas da Planta/Projeto.")
                return False

            # Modo Híbrido: validar fonte dos arquivos
            # Precisa ter OU prenotação (para buscar automaticamente) OU arquivos manuais
            if not tem_arquivo_incra_manual and not tem_arquivo_projeto_manual:
                # Se não tem arquivos manuais, precisa de prenotação
                if not self.numero_prenotacao.get():
                    messagebox.showerror(
                        "Erro",
                        "Por favor:\n\n"
                        "• Insira o Número de Prenotação (para busca automática)\n"
                        "OU\n"
                        "• Selecione os arquivos manualmente"
                    )
                    return False
            else:
                # Modo híbrido: validar que se selecionou um arquivo manual, precisa de prenotação para o outro
                if tem_arquivo_incra_manual and not tem_arquivo_projeto_manual:
                    # Tem INCRA manual, precisa de prenotação para buscar Projeto
                    if not self.numero_prenotacao.get():
                        messagebox.showerror(
                            "Erro",
                            "Você selecionou o arquivo INCRA manualmente.\n\n"
                            "Por favor:\n"
                            "• Insira o Número de Prenotação (para buscar o Projeto)\n"
                            "OU\n"
                            "• Selecione o arquivo do Projeto manualmente"
                        )
                        return False

                if tem_arquivo_projeto_manual and not tem_arquivo_incra_manual:
                    # Tem Projeto manual, precisa de prenotação para buscar INCRA
                    if not self.numero_prenotacao.get():
                        messagebox.showerror(
                            "Erro",
                            "Você selecionou o arquivo do Projeto manualmente.\n\n"
                            "Por favor:\n"
                            "• Insira o Número de Prenotação (para buscar o INCRA)\n"
                            "OU\n"
                            "• Selecione o arquivo do INCRA manualmente"
                        )
                        return False

        return True

    def _buscar_arquivo_tiff(self) -> Optional[str]:
        """Busca arquivo TIFF na rede baseado no número de prenotação."""
        numero = int(self.numero_prenotacao.get())
        numero_formatado = f"{numero:08d}"

        # Calcular subpasta
        milhar = math.ceil(numero / 1000) * 1000
        subpasta_formatada = f"{milhar:08d}"

        # Montar caminho
        base_path = Path(r"\\192.168.20.100\trabalho\TRABALHO\IMAGENS\IMOVEIS\DOCUMENTOS - DIVERSOS")
        tiff_path = base_path / subpasta_formatada / f"{numero_formatado}.tif"

        self._atualizar_status(f"🔍 Buscando: {tiff_path}")

        if tiff_path.exists():
            return str(tiff_path)

        return None

    def _converter_tiff_para_pdf(self, tiff_path: str) -> str:
        """Copia TIFF para Downloads e converte para PDF."""
        downloads_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        downloads_dir.mkdir(parents=True, exist_ok=True)

        # Copiar TIFF
        tiff_filename = Path(tiff_path).name
        tiff_dest = downloads_dir / tiff_filename
        shutil.copy2(tiff_path, tiff_dest)

        # Converter para PDF
        pdf_path = downloads_dir / f"{Path(tiff_filename).stem}.pdf"

        # Abrir TIFF multi-página
        img = Image.open(tiff_dest)
        images = []

        try:
            while True:
                images.append(img.copy().convert('RGB'))
                img.seek(img.tell() + 1)
        except EOFError:
            pass

        # Salvar como PDF
        if images:
            images[0].save(
                pdf_path,
                save_all=True,
                append_images=images[1:],
                resolution=200.0
            )

        return str(pdf_path)

    def _extrair_paginas_manual(self, pdf_path: str, paginas_str: str, output_filename: str) -> str:
        """
        Extrai páginas específicas do PDF baseado numa string de números.

        Args:
            pdf_path: Caminho do PDF original
            paginas_str: String com números de páginas separados por vírgula (ex: "1,2,4,7")
            output_filename: Nome do arquivo de saída

        Returns:
            Caminho do PDF extraído
        """
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_pdf = output_dir / output_filename

        # Parsear a string de páginas
        try:
            # Remover espaços e split por vírgula
            paginas_str = paginas_str.strip()
            if not paginas_str:
                raise ValueError("Nenhuma página especificada")

            # Converter para lista de inteiros (páginas começam em 0 no PyPDF2)
            paginas_lista = [int(p.strip()) - 1 for p in paginas_str.split(',')]

            if not paginas_lista:
                raise ValueError("Nenhuma página válida especificada")

        except ValueError as e:
            print(f"❌ Erro ao parsear páginas: {e}")
            messagebox.showerror(
                "Erro",
                f"Formato de páginas inválido!\n\nUse números separados por vírgula.\nExemplo: 1,2,4,7"
            )
            return ""

        # Abrir o PDF e extrair as páginas
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()

                total_paginas = len(reader.pages)

                # Validar e adicionar cada página
                for page_num in paginas_lista:
                    if 0 <= page_num < total_paginas:
                        writer.add_page(reader.pages[page_num])
                        print(f"✅ Página {page_num + 1} adicionada")
                    else:
                        print(f"⚠️ Página {page_num + 1} não existe no PDF (total: {total_paginas})")

                # Salvar PDF extraído
                with open(output_pdf, 'wb') as output_file:
                    writer.write(output_file)

                print(f"✅ PDF criado com {len(writer.pages)} página(s): {output_pdf}")
                return str(output_pdf)

        except Exception as e:
            print(f"❌ Erro ao extrair páginas: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao extrair páginas do PDF:\n{e}")
            return ""

    def _extrair_memorial_incra_do_pdf(self, pdf_path: str) -> str:
        """Extrai páginas do Memorial INCRA do PDF usando IA."""
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_pdf = output_dir / "memorial_incra_extraido.pdf"

        # Usar Gemini 2.0 Flash Experimental
        api_key = self.config_manager.get_api_key()
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash-lite')

        images = convert_from_path(pdf_path, dpi=150, poppler_path=POPPLER_PATH)
        paginas_encontradas = []

        for i, img in enumerate(images):
            temp_img_path = output_dir / f"temp_page_{i}.jpg"
            img.save(temp_img_path, 'JPEG')

            prompt = """
            Analise esta imagem e responda apenas com 'SIM' ou 'NAO':
            Esta página contém o Memorial Descritivo do INCRA?

            CRITÉRIOS DE IDENTIFICAÇÃO (Todos devem estar presentes):

            1. CABEÇALHO OFICIAL DO INCRA (Topo da Página):
               - Deve conter: "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO"
               - Logo abaixo: "INSTITUTO NACIONAL DE COLONIZAÇÃO E REFORMA AGRÁRIA"
               - Título em destaque: "MEMORIAL DESCRITIVO"
               - Dados do Imóvel: "Denominação:", "Proprietário(a):", "Matrícula do imóvel:", "Município/UF:"

            2. MARCADOR DE INÍCIO DA TABELA (Gatilho/Anchor):
               - Procure pelo subtítulo em MAIÚSCULAS: "DESCRIÇÃO DA PARCELA"
               - Imediatamente abaixo, deve haver cabeçalho da tabela com as colunas:
                 "VÉRTICE", "Longitude", "Latitude", "SEGMENTO VANTE"

            IMPORTANTE: A página deve conter TANTO o cabeçalho oficial quanto o marcador de início da tabela.

            Responda apenas: SIM ou NAO
            """

            img_upload = None
            tentativas = 0
            max_tentativas = 3

            while tentativas < max_tentativas:
                try:
                    img_upload = Image.open(temp_img_path)
                    response = model.generate_content([prompt, img_upload])
                    resposta = response.text.strip().upper()

                    if 'SIM' in resposta:
                        paginas_encontradas.append(i)
                    break  # Sucesso, sair do loop

                except Exception as e:
                    erro_str = str(e).lower()
                    # Detectar erros de rate limit
                    if '429' in erro_str or 'quota' in erro_str or 'rate limit' in erro_str or 'resource exhausted' in erro_str or 'resource has been exhausted' in erro_str:
                        tentativas += 1
                        tempo_espera = 60

                        if tentativas < max_tentativas:
                            # Atualizar pop-up informando o usuário
                            if hasattr(self, 'progress_window') and self.progress_window:
                                self.root.after(0, lambda: self._atualizar_progresso(
                                    self.progress_bar['value'],
                                    "⏸️ LIMITE DE API ATINGIDO",
                                    f"Aguardando {tempo_espera}s para continuar... (Tentativa {tentativas}/{max_tentativas})"
                                ))

                            print(f"⚠️ Limite de API atingido. Pausando por {tempo_espera} segundos... (Tentativa {tentativas}/{max_tentativas})")
                            time.sleep(tempo_espera)
                        else:
                            print(f"❌ Erro: Limite de API excedido após {max_tentativas} tentativas na página {i}")
                            break
                    else:
                        # Outro tipo de erro
                        print(f"Erro ao analisar página {i}: {e}")
                        break
                finally:
                    # Fechar a imagem antes de deletar o arquivo
                    if img_upload:
                        img_upload.close()
                        img_upload = None

            # Deletar arquivo temporário com tratamento de erro
            try:
                temp_img_path.unlink()
            except Exception as e:
                print(f"Aviso: Não foi possível deletar {temp_img_path}: {e}")

        # Extrair páginas
        if paginas_encontradas:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()

                for page_num in paginas_encontradas:
                    writer.add_page(reader.pages[page_num])

                with open(output_pdf, 'wb') as output_file:
                    writer.write(output_file)

        return str(output_pdf)

    def _extrair_projeto_do_pdf(self, pdf_path: str) -> str:
        """Extrai páginas da Planta/Projeto do PDF."""
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_pdf = output_dir / "projeto_extraido.pdf"

        # Usar Gemini 2.0 Flash Experimental
        api_key = self.config_manager.get_api_key()
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash-lite')

        images = convert_from_path(pdf_path, dpi=150, poppler_path=POPPLER_PATH)
        paginas_encontradas = []

        for i, img in enumerate(images):
            temp_img_path = output_dir / f"temp_page_{i}.jpg"
            img.save(temp_img_path, 'JPEG')

            prompt = """
            Analise esta imagem e responda apenas com 'SIM' ou 'NAO':
            Esta página contém a Planta/Projeto de Georreferenciamento?

            CRITÉRIOS DE IDENTIFICAÇÃO DA PLANTA TÉCNICA:

            1. PALAVRAS-CHAVE VISUAIS (no topo ou laterais):
               - "PLANTA DE SITUAÇÃO" ou "PLANTA DO IMÓVEL" ou "PLANTA DO IMÓVEL GEORREFERENCIADO"

            2. ELEMENTO DE MAPA:
               - Referências a provedores de mapas: palavra "Google" (geralmente no canto inferior)
               - Ou texto "Imagem de Satélite"
               - Desenho técnico com linhas vetoriais representando um polígono (terreno)

            3. TABELA DE COORDENADAS (Busca Flexível - pode estar na esquerda, direita ou embaixo):
               - Cabeçalho da tabela deve conter palavras próximas:
                 * "VÉRTICE" (ou "Vért", "Pt")
                 * "AZIMUTE" (ou "Azim")
                 * "DISTÂNCIA" (ou "Dist", "Dist. (m)")
                 * "COORDENADAS" (ou "Latitude/Longitude" ou "N/E")
               - Título da tabela pode ser: "Tabela de Coordenadas", "Memorial Analítico", "Dados da Poligonal"

            4. CRITÉRIO DE PARADA (para não confundir com outras seções):
               - NÃO deve ser página de "CONVENÇÕES", "DECLARAÇÃO", "ASSINATURAS" ou "ESCALA"

            IMPORTANTE: A página deve ser uma Planta Técnica com mapa E tabela de coordenadas.

            Responda apenas: SIM ou NAO
            """

            img_upload = None
            tentativas = 0
            max_tentativas = 3

            while tentativas < max_tentativas:
                try:
                    img_upload = Image.open(temp_img_path)
                    response = model.generate_content([prompt, img_upload])
                    resposta = response.text.strip().upper()

                    if 'SIM' in resposta:
                        paginas_encontradas.append(i)
                    break  # Sucesso, sair do loop

                except Exception as e:
                    erro_str = str(e).lower()
                    # Detectar erros de rate limit
                    if '429' in erro_str or 'quota' in erro_str or 'rate limit' in erro_str or 'resource exhausted' in erro_str or 'resource has been exhausted' in erro_str:
                        tentativas += 1
                        tempo_espera = 60

                        if tentativas < max_tentativas:
                            # Atualizar pop-up informando o usuário
                            if hasattr(self, 'progress_window') and self.progress_window:
                                self.root.after(0, lambda: self._atualizar_progresso(
                                    self.progress_bar['value'],
                                    "⏸️ LIMITE DE API ATINGIDO",
                                    f"Aguardando {tempo_espera}s para continuar... (Tentativa {tentativas}/{max_tentativas})"
                                ))

                            print(f"⚠️ Limite de API atingido. Pausando por {tempo_espera} segundos... (Tentativa {tentativas}/{max_tentativas})")
                            time.sleep(tempo_espera)
                        else:
                            print(f"❌ Erro: Limite de API excedido após {max_tentativas} tentativas na página {i}")
                            break
                    else:
                        # Outro tipo de erro
                        print(f"Erro ao analisar página {i}: {e}")
                        break
                finally:
                    # Fechar a imagem antes de deletar o arquivo
                    if img_upload:
                        img_upload.close()
                        img_upload = None

            # Deletar arquivo temporário com tratamento de erro
            try:
                temp_img_path.unlink()
            except Exception as e:
                print(f"Aviso: Não foi possível deletar {temp_img_path}: {e}")

        # Extrair páginas
        if paginas_encontradas:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()

                for page_num in paginas_encontradas:
                    writer.add_page(reader.pages[page_num])

                with open(output_pdf, 'wb') as output_file:
                    writer.write(output_file)

        return str(output_pdf)

    def _salvar_backups_pdfs(self):
        """Salva backups dos PDFs extraídos."""
        try:
            # Tentar várias localizações comuns
            possible_docs = [
                Path.home() / "Documents",  # Inglês/Linux
                Path.home() / "Documentos",  # Português
                Path.home()  # Fallback
            ]

            docs_dir = None
            for path in possible_docs:
                if path.exists() and path.is_dir():
                    docs_dir = path / "Relatórios INCRA"
                    break

            if docs_dir is None:
                docs_dir = Path.home() / "Relatórios INCRA"

            # Criar diretório principal
            docs_dir.mkdir(parents=True, exist_ok=True)
            print(f"📁 Diretório principal criado/verificado: {docs_dir}")

            # Criar subdiretórios
            incra_dir = docs_dir / "PDF_INCRAS"
            projeto_dir = docs_dir / "PDF_PLANTAS"

            incra_dir.mkdir(parents=True, exist_ok=True)
            projeto_dir.mkdir(parents=True, exist_ok=True)

            print(f"📁 Pasta PDF_INCRAS criada em: {incra_dir}")
            print(f"📁 Pasta PDF_PLANTAS criada em: {projeto_dir}")

            numero = self.numero_prenotacao.get()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            if self.pdf_extraido_incra and Path(self.pdf_extraido_incra).exists():
                dest_incra = incra_dir / f"INCRA_{numero}_{timestamp}.pdf"
                shutil.copy2(self.pdf_extraido_incra, dest_incra)
                print(f"✅ PDF INCRA salvo em: {dest_incra}")
            else:
                print(f"⚠️ PDF INCRA não encontrado ou não existe")

            if self.pdf_extraido_projeto and Path(self.pdf_extraido_projeto).exists():
                dest_projeto = projeto_dir / f"PROJETO_{numero}_{timestamp}.pdf"
                shutil.copy2(self.pdf_extraido_projeto, dest_projeto)
                print(f"✅ PDF PROJETO salvo em: {dest_projeto}")
            else:
                print(f"⚠️ PDF PROJETO não encontrado ou não existe")

        except Exception as e:
            print(f"❌ Erro ao salvar backups: {str(e)}")
            import traceback
            traceback.print_exc()

    def _ao_fechar_aplicacao(self):
        """Executado quando o usuário fecha a aplicação - limpa arquivos temporários."""
        try:
            # Deletar pasta temporária conferencia_geo_temp
            temp_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
                print(f"🗑️ Pasta temporária deletada: {temp_dir}")
        except Exception as e:
            print(f"⚠️ Erro ao deletar pasta temporária: {e}")
        finally:
            # Fechar a aplicação
            self.root.destroy()

    def _limpar_arquivos_backup(self):
        """Limpa todos os arquivos das pastas de backup PDF_INCRAS e PDF_PLANTAS."""
        # Confirmar com o usuário
        resposta = messagebox.askyesno(
            "Confirmar Limpeza",
            "Tem certeza que deseja deletar TODOS os arquivos de backup?\n\n"
            "Isso irá remover todos os PDFs salvos nas pastas:\n"
            "• PDF_INCRAS\n"
            "• PDF_PLANTAS\n\n"
            "Esta ação não pode ser desfeita!",
            icon='warning'
        )

        if not resposta:
            return

        try:
            # Tentar várias localizações comuns
            possible_docs = [
                Path.home() / "Documents",  # Inglês/Linux
                Path.home() / "Documentos",  # Português
                Path.home()  # Fallback
            ]

            docs_dir = None
            for path in possible_docs:
                relatorios_dir = path / "Relatórios INCRA"
                if relatorios_dir.exists() and relatorios_dir.is_dir():
                    docs_dir = relatorios_dir
                    break

            if docs_dir is None:
                docs_dir = Path.home() / "Relatórios INCRA"

            # Pastas a limpar
            incra_dir = docs_dir / "PDF_INCRAS"
            projeto_dir = docs_dir / "PDF_PLANTAS"

            arquivos_deletados = 0

            # Deletar arquivos da pasta INCRA
            if incra_dir.exists():
                for arquivo in incra_dir.glob("*.pdf"):
                    try:
                        arquivo.unlink()
                        arquivos_deletados += 1
                        print(f"🗑️ Deletado: {arquivo.name}")
                    except Exception as e:
                        print(f"⚠️ Erro ao deletar {arquivo.name}: {e}")

            # Deletar arquivos da pasta PROJETO
            if projeto_dir.exists():
                for arquivo in projeto_dir.glob("*.pdf"):
                    try:
                        arquivo.unlink()
                        arquivos_deletados += 1
                        print(f"🗑️ Deletado: {arquivo.name}")
                    except Exception as e:
                        print(f"⚠️ Erro ao deletar {arquivo.name}: {e}")

            # Atualizar status
            if arquivos_deletados > 0:
                self.backup_status_label.config(
                    text=f"✅ {arquivos_deletados} arquivo(s) deletado(s)",
                    fg=self.colors['success']
                )
                messagebox.showinfo(
                    "Limpeza Concluída",
                    f"✅ {arquivos_deletados} arquivo(s) de backup foram deletados com sucesso!"
                )
            else:
                self.backup_status_label.config(
                    text="ℹ️ Nenhum arquivo encontrado",
                    fg=self.colors['info']
                )
                messagebox.showinfo(
                    "Limpeza Concluída",
                    "ℹ️ Nenhum arquivo de backup foi encontrado nas pastas."
                )

            # Limpar status após 5 segundos
            self.root.after(5000, lambda: self.backup_status_label.config(text=""))

        except Exception as e:
            print(f"❌ Erro ao limpar backups: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"❌ Erro ao limpar backups:\n{str(e)}")

    def _limpar_dados(self):
        """Limpa todos os dados para permitir uma nova conferência."""
        # Confirmar com o usuário
        resposta = messagebox.askyesno(
            "Confirmar Limpeza",
            "Deseja limpar todos os dados e fazer uma nova conferência?\n\n"
            "Isso irá resetar:\n"
            "• Todos os campos de entrada\n"
            "• Arquivos selecionados\n"
            "• Dados extraídos\n"
            "• Relatório de comparação\n\n"
            "Os arquivos de backup permanecerão salvos.",
            icon='question'
        )

        if not resposta:
            return

        try:
            # Limpar campos de entrada - Modo Automático
            self.numero_prenotacao.set("")
            self.paginas_incra_auto.set("")
            self.paginas_projeto_auto.set("")
            self.arquivo_manual_incra_auto.set("")
            self.arquivo_manual_projeto_auto.set("")

            # Limpar campos de entrada - Modo Manual
            self.incra_path.set("")
            self.projeto_path.set("")
            self.paginas_incra_manual.set("")
            self.paginas_projeto_manual.set("")

            # Resetar variáveis de dados extraídos
            self.incra_excel_path = None
            self.projeto_excel_path = None
            self.incra_data = None
            self.projeto_data = None
            self.pdf_extraido_incra = None
            self.pdf_extraido_projeto = None

            # Limpar área de resultados
            self.resultado_text.delete(1.0, tk.END)
            self.resultado_text.insert(1.0, "Pronto para uma nova conferência...")

            # Esconder preview frame (se existir)
            if hasattr(self, 'preview_frame'):
                self.preview_frame.pack_forget()

            # Resetar status
            self._atualizar_status("✨ Pronto para iniciar nova conferência")

            # Resetar sub-modos para padrão
            self.modo_automatico_tipo.set("paginas")
            self.modo_manual_tipo.set("completo")

            # Atualizar visuals dos sub-modos
            if hasattr(self, '_atualizar_visual_submodo_auto'):
                self._atualizar_visual_submodo_auto()
            if hasattr(self, '_atualizar_visual_submodo_manual'):
                self._atualizar_visual_submodo_manual()

            # Esconder campos de páginas
            if hasattr(self, 'paginas_frame_auto'):
                self.paginas_frame_auto.pack_forget()
            if hasattr(self, 'paginas_frame_manual'):
                self.paginas_frame_manual.pack_forget()

            # Atualizar modo para refletir estado inicial
            self._alternar_modo_automatico()
            self._alternar_modo_manual()

            print("✅ Dados limpos com sucesso! Pronto para nova conferência.")

        except Exception as e:
            print(f"❌ Erro ao limpar dados: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"❌ Erro ao limpar dados:\n{str(e)}")

    def _gerar_previews(self):
        """Gera thumbnails dos documentos extraídos."""
        if self.pdf_extraido_incra and Path(self.pdf_extraido_incra).exists():
            images = convert_from_path(self.pdf_extraido_incra, dpi=100, first_page=1, last_page=1, poppler_path=POPPLER_PATH)
            if images:
                self.preview_incra_image = images[0]
                self.preview_incra_image.thumbnail((300, 400))

                photo = ImageTk.PhotoImage(self.preview_incra_image)
                self.incra_preview_label.config(image=photo)
                self.incra_preview_label.image = photo

        if self.pdf_extraido_projeto and Path(self.pdf_extraido_projeto).exists():
            images = convert_from_path(self.pdf_extraido_projeto, dpi=100, first_page=1, last_page=1, poppler_path=POPPLER_PATH)
            if images:
                self.preview_projeto_image = images[0]
                self.preview_projeto_image.thumbnail((300, 400))

                photo = ImageTk.PhotoImage(self.preview_projeto_image)
                self.projeto_preview_label.config(image=photo)
                self.projeto_preview_label.image = photo

    def _confirmar_documentos_automaticos(self):
        """Usuário confirmou documentos - prosseguir com comparação."""
        self.incra_path.set(self.pdf_extraido_incra)
        self.projeto_path.set(self.pdf_extraido_projeto)

        self.preview_frame.pack_forget()

        self._comparar_manual()

    def _alternar_para_manual(self):
        """Usuário optou por fazer manual."""
        self.preview_frame.pack_forget()
        self._selecionar_modo("manual")
        self._habilitar_botoes()
        messagebox.showinfo(
            "Modo Manual",
            "Selecione manualmente os arquivos corretos no Modo Manual."
        )

    # ========== EXTRAÇÃO E COMPARAÇÃO ==========

    def _extrair_pdf_para_excel(self, pdf_path: str, tipo: str = "normal") -> tuple[str, Dict]:
        """Extrai dados de um PDF memorial para Excel."""
        try:
            api_key = self.config_manager.get_api_key()
            genai.configure(api_key=api_key)

            output_dir = Path(tempfile.gettempdir()) / "conferencia_geo"
            output_dir.mkdir(parents=True, exist_ok=True)

            if not output_dir.exists():
                raise RuntimeError(f"Não foi possível criar o diretório: {output_dir}")

            nome_base = Path(pdf_path).stem
            excel_path = output_dir / f"{nome_base}_extraido.xlsx"

            if tipo == "incra":
                dados = extrair_memorial_incra(pdf_path, api_key)
            else:
                dados = extract_table_from_pdf(pdf_path, api_key)

            if not dados or 'data' not in dados:
                raise ValueError("Nenhum dado foi extraído do PDF")

            create_excel_file(dados, str(excel_path))

            if not excel_path.exists():
                raise RuntimeError(f"Arquivo Excel não foi criado")

            if excel_path.stat().st_size == 0:
                raise RuntimeError(f"Arquivo Excel está vazio")

            return str(excel_path), dados

        except Exception as e:
            error_msg = f"Erro ao extrair PDF para Excel: {str(e)}"
            raise RuntimeError(error_msg) from e

    def _normalizar_coordenada(self, coord: str) -> str:
        """Normaliza coordenadas para comparação."""
        if not coord:
            return ""

        coord = str(coord).strip()
        coord = coord.replace("′", "'").replace("″", '"')

        if coord.startswith("-"):
            coord = coord[1:].strip()

        coord = coord.replace(" W", "").replace(" S", "").strip()
        coord = coord.strip().strip('"').strip("'").strip()

        return coord

    def _limpar_string(self, valor) -> str:
        """Limpa strings e converte pontos para vírgulas."""
        if valor is None:
            return ""

        valor_limpo = str(valor).strip()

        while "  " in valor_limpo:
            valor_limpo = valor_limpo.replace("  ", " ")

        valor_limpo = valor_limpo.replace(".", ",")

        return valor_limpo

    def _construir_relatorio_comparacao(self, incluir_projeto: bool, incluir_memorial: bool) -> str:
        """Constrói relatório HTML comparando dados estruturados."""
        html = []

        html.append("""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório de Conferência INCRA</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            background: #f5f5f5;
            padding: 20px;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border: 1px solid #e0e0e0;
        }
        h1 {
            color: #1a1a1a;
            text-align: center;
            margin-bottom: 10px;
            font-size: 28px;
            font-weight: 600;
        }
        .subtitle {
            text-align: center;
            color: #666;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .info-box {
            background: #f8f9fa;
            padding: 15px 20px;
            border-radius: 6px;
            margin-bottom: 30px;
            border-left: 4px solid #2c5282;
        }
        .info-box strong {
            color: #1a1a1a;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            border: 1px solid #e0e0e0;
        }
        th {
            background: #2c5282;
            color: white;
            padding: 14px 15px;
            text-align: left;
            font-weight: 600;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        td {
            padding: 12px 15px;
            border-bottom: 1px solid #e8e8e8;
            font-size: 13px;
            color: #333;
        }
        tr:hover {
            background-color: #f8f9fa;
        }
        .identico {
            background-color: #e8f5e9 !important;
            border-left: 4px solid #2e7d32;
        }
        .diferente {
            background-color: #ffebee !important;
            border-left: 4px solid #c62828;
            font-weight: 600;
        }
        .resumo {
            background: #2c5282;
            color: white;
            padding: 25px 30px;
            border-radius: 6px;
            margin-top: 30px;
        }
        .resumo h2 {
            margin-bottom: 20px;
            font-size: 22px;
            font-weight: 600;
        }
        .resumo h4 {
            margin-top: 18px;
            margin-bottom: 10px;
            font-size: 16px;
            font-weight: 600;
            border-bottom: 1px solid rgba(255,255,255,0.3);
            padding-bottom: 8px;
        }
        .resumo p {
            margin: 8px 0;
            font-size: 15px;
        }
        .section-title {
            color: #1a1a1a;
            margin: 40px 0 20px 0;
            padding-bottom: 10px;
            border-bottom: 3px solid #2c5282;
            font-size: 22px;
            font-weight: 600;
        }
        .rodape {
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #e0e0e0;
            color: #888;
            font-size: 12px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📋 RELATÓRIO DE CONFERÊNCIA INCRA</h1>
        <p class="subtitle">Sistema Profissional de Análise e Verificação v4.0</p>
""")

        html.append(f"""
        <div class="info-box">
            <p><strong>📅 Data:</strong> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
            <p><strong>📋 Nº Prenotação:</strong> {self.numero_prenotacao.get()}</p>
        </div>
""")

        # Carregar dados
        wb_incra = load_workbook(self.incra_excel_path)
        ws_incra = wb_incra.active
        dados_incra = list(ws_incra.iter_rows(values_only=True))

        wb_projeto = load_workbook(self.projeto_excel_path)
        ws_projeto = wb_projeto.active
        dados_projeto = list(ws_projeto.iter_rows(values_only=True))

        identicos_vertice = 0
        diferencas_vertice = 0
        identicos_segmento = 0
        diferencas_segmento = 0
        vertices_com_diferenca = []

        # VÉRTICE - Formato reorganizado e mais claro
        html.append('<h2 class="section-title">📐 COMPARAÇÃO DE VÉRTICES</h2>')
        html.append('<p style="color: #666; margin-bottom: 20px;">Cada bloco representa um vértice completo com todas as suas coordenadas</p>')
        html.append('<table>')
        html.append('<thead><tr>')
        html.append('<th style="width: 80px;">Vértice</th>')
        html.append('<th style="width: 150px;">Coluna</th>')
        html.append('<th style="width: 35%;">INCRA</th>')
        html.append('<th style="width: 35%;">PROJETO</th>')
        html.append('<th style="width: 120px;">Status</th>')
        html.append('</tr></thead><tbody>')

        max_rows = max(len(dados_incra), len(dados_projeto))

        for i in range(1, max_rows):
            incra_row = dados_incra[i] if i < len(dados_incra) else []
            projeto_row = dados_projeto[i] if i < len(dados_projeto) else []

            codigo_incra = self._limpar_string(incra_row[0] if len(incra_row) > 0 else "")
            codigo_projeto = self._limpar_string(projeto_row[0] if len(projeto_row) > 0 else "")

            long_incra = self._normalizar_coordenada(self._limpar_string(incra_row[1] if len(incra_row) > 1 else ""))
            long_projeto = self._normalizar_coordenada(self._limpar_string(projeto_row[1] if len(projeto_row) > 1 else ""))

            lat_incra = self._normalizar_coordenada(self._limpar_string(incra_row[2] if len(incra_row) > 2 else ""))
            lat_projeto = self._normalizar_coordenada(self._limpar_string(projeto_row[2] if len(projeto_row) > 2 else ""))

            alt_incra = self._limpar_string(incra_row[3] if len(incra_row) > 3 else "")
            alt_projeto = self._limpar_string(projeto_row[3] if len(projeto_row) > 3 else "")

            # Lista de campos para este vértice
            campos = [
                ("CÓDIGO", codigo_incra, codigo_projeto, "Col A"),
                ("LONGITUDE", long_incra, long_projeto, "Col B"),
                ("LATITUDE", lat_incra, lat_projeto, "Col C"),
                ("ALTITUDE", alt_incra, alt_projeto, "Col D")
            ]

            vertice_tem_diferenca = False

            # Adicionar linha de separação visual entre vértices
            if i > 1:
                html.append('<tr style="height: 3px; background: #2c5282;"><td colspan="5"></td></tr>')

            # Iterar pelos campos deste vértice
            for idx, (campo, val_incra, val_projeto, col_name) in enumerate(campos):
                status_classe = "identico" if val_incra == val_projeto else "diferente"
                status_texto = "✅ Idêntico" if val_incra == val_projeto else "❌ Diferente"

                if val_incra == val_projeto:
                    identicos_vertice += 1
                else:
                    diferencas_vertice += 1
                    vertice_tem_diferenca = True

                html.append(f'<tr class="{status_classe}">')

                # Mostrar número do vértice apenas na primeira linha
                if idx == 0:
                    html.append(f'<td rowspan="4" style="text-align: center; font-size: 18px; font-weight: bold; background: #f0f0f0; border-right: 3px solid #2c5282;">#{i}</td>')

                html.append(f'<td><strong>{campo}</strong><br><span style="font-size: 11px; color: #999;">{col_name}</span></td>')
                html.append(f'<td>{val_incra}</td>')
                html.append(f'<td>{val_projeto}</td>')
                html.append(f'<td style="text-align: center;">{status_texto}</td>')
                html.append('</tr>')

            if vertice_tem_diferenca:
                vertices_com_diferenca.append(i)

        html.append('</tbody></table>')

        # SEGMENTO VANTE - Formato reorganizado e mais claro
        segmentos_com_diferenca = []
        html.append('<h2 class="section-title">🔄 COMPARAÇÃO DE SEGMENTOS VANTE</h2>')
        html.append('<p style="color: #666; margin-bottom: 20px;">Cada bloco representa um segmento completo com todas as suas medidas</p>')
        html.append('<table>')
        html.append('<thead><tr>')
        html.append('<th style="width: 80px;">Segmento</th>')
        html.append('<th style="width: 150px;">Coluna</th>')
        html.append('<th style="width: 35%;">INCRA</th>')
        html.append('<th style="width: 35%;">PROJETO</th>')
        html.append('<th style="width: 120px;">Status</th>')
        html.append('</tr></thead><tbody>')

        for i in range(1, max_rows):
            incra_row = dados_incra[i] if i < len(dados_incra) else []
            projeto_row = dados_projeto[i] if i < len(dados_projeto) else []

            cod_seg_incra = self._limpar_string(incra_row[4] if len(incra_row) > 4 else "")
            cod_seg_projeto = self._limpar_string(projeto_row[4] if len(projeto_row) > 4 else "")

            azim_incra = self._limpar_string(incra_row[5] if len(incra_row) > 5 else "")
            azim_projeto = self._limpar_string(projeto_row[5] if len(projeto_row) > 5 else "")

            dist_incra = self._limpar_string(incra_row[6] if len(incra_row) > 6 else "")
            dist_projeto = self._limpar_string(projeto_row[6] if len(projeto_row) > 6 else "")

            campos = [
                ("CÓDIGO", cod_seg_incra, cod_seg_projeto, "Col E"),
                ("AZIMUTE", azim_incra, azim_projeto, "Col F"),
                ("DISTÂNCIA", dist_incra, dist_projeto, "Col G")
            ]

            segmento_tem_diferenca = False

            # Adicionar linha de separação visual entre segmentos
            if i > 1:
                html.append('<tr style="height: 3px; background: #2c5282;"><td colspan="5"></td></tr>')

            for idx, (campo, val_incra, val_projeto, col_name) in enumerate(campos):
                status_classe = "identico" if val_incra == val_projeto else "diferente"
                status_texto = "✅ Idêntico" if val_incra == val_projeto else "❌ Diferente"

                if val_incra == val_projeto:
                    identicos_segmento += 1
                else:
                    diferencas_segmento += 1
                    segmento_tem_diferenca = True

                html.append(f'<tr class="{status_classe}">')

                # Mostrar número do segmento apenas na primeira linha
                if idx == 0:
                    html.append(f'<td rowspan="3" style="text-align: center; font-size: 18px; font-weight: bold; background: #f0f0f0; border-right: 3px solid #2c5282;">#{i}</td>')

                html.append(f'<td><strong>{campo}</strong><br><span style="font-size: 11px; color: #999;">{col_name}</span></td>')
                html.append(f'<td>{val_incra}</td>')
                html.append(f'<td>{val_projeto}</td>')
                html.append(f'<td style="text-align: center;">{status_texto}</td>')
                html.append('</tr>')

            if segmento_tem_diferenca:
                segmentos_com_diferenca.append(i)

        html.append('</tbody></table>')

        # RESUMO
        identicos_total = identicos_vertice + identicos_segmento
        diferencas_total = diferencas_vertice + diferencas_segmento

        # Criar lista de vértices e segmentos com diferenças
        vertices_str = ", ".join([f"#{v}" for v in vertices_com_diferenca]) if vertices_com_diferenca else "Nenhum"
        segmentos_str = ", ".join([f"#{s}" for s in segmentos_com_diferenca]) if segmentos_com_diferenca else "Nenhum"

        html.append(f"""
        <div class="resumo">
            <h2>📊 RESUMO DA COMPARAÇÃO</h2>

            <h4>📍 VÉRTICES:</h4>
            <p>✅ Campos idênticos: <strong>{identicos_vertice}</strong></p>
            <p>❌ Campos diferentes: <strong>{diferencas_vertice}</strong></p>
            <p>⚠️ Vértices com diferenças: <strong>{vertices_str}</strong></p>

            <h4>🔄 SEGMENTOS VANTE:</h4>
            <p>✅ Campos idênticos: <strong>{identicos_segmento}</strong></p>
            <p>❌ Campos diferentes: <strong>{diferencas_segmento}</strong></p>
            <p>⚠️ Segmentos com diferenças: <strong>{segmentos_str}</strong></p>

            <h4>🎯 TOTAL GERAL:</h4>
            <p>✅ Total de campos idênticos: <strong>{identicos_total}</strong></p>
            <p>❌ Total de campos diferentes: <strong>{diferencas_total}</strong></p>
            <p>📋 Total de vértices analisados: <strong>{max_rows - 1}</strong></p>
        </div>
        <div class="rodape">
            <p>Relatório gerado automaticamente pelo Sistema de Verificação INCRA v4.0</p>
        </div>
    </div>
</body>
</html>
""")

        return "".join(html)

    def _salvar_e_abrir_relatorio(self, conteudo_html: str):
        """Salva relatório automaticamente e abre no navegador."""
        # Tentar várias localizações comuns
        possible_docs = [
            Path.home() / "Documents",  # Inglês/Linux
            Path.home() / "Documentos",  # Português
            Path.home()  # Fallback
        ]

        relatorios_dir = None
        for path in possible_docs:
            if path.exists():
                relatorios_dir = path / "Relatórios INCRA"
                break

        if relatorios_dir is None:
            relatorios_dir = Path.home() / "Relatórios INCRA"

        relatorios_dir.mkdir(parents=True, exist_ok=True)

        numero = self.numero_prenotacao.get()
        nome_arquivo = f"Relatório_INCRA_{numero}.html"
        caminho_completo = relatorios_dir / nome_arquivo

        with open(caminho_completo, 'w', encoding='utf-8') as f:
            f.write(conteudo_html)

        webbrowser.open(f'file://{caminho_completo}')

        self._atualizar_status(f"✅ Relatório salvo: {caminho_completo}")
        print(f"✅ Relatório HTML salvo em: {caminho_completo}")

    def _mostrar_resumo_no_texto(self):
        """Mostra resumo simplificado na área de texto."""
        self.resultado_text.delete(1.0, tk.END)

        resumo = f"""
╔════════════════════════════════════════════════════════════════╗
║          COMPARAÇÃO CONCLUÍDA COM SUCESSO                      ║
╚════════════════════════════════════════════════════════════════╝

📋 Número de Prenotação: {self.numero_prenotacao.get()}
📅 Data: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}

✅ O relatório HTML completo foi gerado e aberto automaticamente.
📁 Local: Documentos\\Relatórios INCRA\\Relatório_INCRA_{self.numero_prenotacao.get()}.html

💡 Consulte o relatório HTML para ver todos os detalhes da comparação.
"""

        self.resultado_text.insert(1.0, resumo)


def main():
    """Função principal."""
    root = tk.Tk()
    app = VerificadorGeorreferenciamento(root)
    root.mainloop()


if __name__ == "__main__":
    main()